# app.py
# Versión estable V5.8 FINAL: conciliación bancaria continua Grupo Dancona.
# Separa saldo Flexxus real, saldo ajustado conciliatorio y total importado .xls.
# Integra login, historial, administrador de archivos, diagnóstico/reset,
# conciliación anterior, pendientes abiertos con ID estable, regularizaciones,
# matching agregado MB-EXT/PAV-QR, Pedidos Ya y exportación Flexxus.
# Streamlit Cloud requirements sugeridos:
# streamlit
# pandas
# numpy
# openpyxl
# xlrd
# xlwt

import io
import re
import uuid
import warnings
import json
import base64
import sys
import traceback
import unicodedata
from datetime import datetime
from typing import Dict, List, Optional, Tuple

import numpy as np
import pandas as pd
import streamlit as st
import xlwt
try:
    import requests
except Exception:
    requests = None
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")

# =============================================================================
# CONFIG
# =============================================================================

st.set_page_config(
    page_title="Conciliación Bancaria Dancona · V5.8",
    page_icon="🏦",
    layout="wide",
    initial_sidebar_state="expanded",
)

LOCAL_MAP = {
    "29841642": "Local 1",
    "29841627": "Local 2",
    "29841683": "Local 3",
    "29841670": "Local 5",
    "29841705": "Local 6",
    "31995644": "Local 7",
    "32032899": "Local 8",
}

IMP_CATS = {
    "IMPUESTO_LEY25413_DEB",
    "IMPUESTO_LEY25413_CRED",
    "RETENCION_IIBB",
    "IVA",
    "GASTO_BANCARIO",
    "DEBITO_LIQ_TARJETA",
    "DEBITO_PAGO_DIRECTO",
}

CAT_LABELS = {
    "QR": "QR / CR DEBIN SPOT",
    "LIQUIDACION_TARJETA": "Liquidación tarjeta",
    "DEBITO_LIQ_TARJETA": "Débito liquidación tarjeta",
    "TRANSFERENCIA_ENTRANTE": "Transferencia entrante",
    "TRANSFERENCIA_SALIENTE": "Transferencia saliente",
    "IMPUESTO_LEY25413_DEB": "Anticipo Ganancias Ley 25.413 débitos",
    "IMPUESTO_LEY25413_CRED": "Anticipo Ganancias Ley 25.413 créditos",
    "RETENCION_IIBB": "Retención Ingresos Brutos Mendoza",
    "IVA": "IVA base / IVA sobre gastos",
    "GASTO_BANCARIO": "Gasto bancario / comisión",
    "DEBITO_PAGO_DIRECTO": "Débito pago directo",
    "OTRO": "Otro",
    "AJUSTE_MENOR_REDONDEO": "Ajuste menor redondeo/extracto",
}

def normalize_category_code(value: str) -> str:
    """Convierte etiquetas visibles del Excel anterior a códigos internos estables."""
    c = norm_txt(value)
    compact = re.sub(r"[^A-Z0-9]", "", c)
    if not c:
        return "OTRO"
    if c in CAT_LABELS:
        return c
    if "CR DEBIN SPOT" in c or compact == "QR" or "QRCRDEBINSPOT" in compact:
        return "QR"
    if "LIQUIDACIONTARJETA" in compact or "TARJETA" in compact:
        return "LIQUIDACION_TARJETA"
    if "DEBITOLIQUIDACIONTARJETA" in compact or "DEBLIQ" in compact or "MASTERCARD" in compact:
        return "DEBITO_LIQ_TARJETA"
    if "TRANSFERENCIAENTRANTE" in compact:
        return "TRANSFERENCIA_ENTRANTE"
    if "TRANSFERENCIASALIENTE" in compact:
        return "TRANSFERENCIA_SALIENTE"
    if "INGRESOSBRUTOS" in compact or "IIBB" in compact or "RETENCIONINGRESOSBRUTOS" in compact:
        return "RETENCION_IIBB"
    if "25413" in compact and ("DEB" in compact or "DEBITO" in compact):
        return "IMPUESTO_LEY25413_DEB"
    if "25413" in compact and ("CRED" in compact or "CREDIT" in compact or "CRE" in compact):
        return "IMPUESTO_LEY25413_CRED"
    if compact.startswith("IVA") or "IVABASE" in compact:
        return "IVA"
    if "GASTO" in compact or "COMTRANSFE" in compact or "COMISION" in compact:
        return "GASTO_BANCARIO"
    if "PAGODIRECTO" in compact:
        return "DEBITO_PAGO_DIRECTO"
    if "AJUSTEMENOR" in compact or "REDONDEO" in compact:
        return "AJUSTE_MENOR_REDONDEO"
    if "PAV" in compact:
        return "PAV"
    return str(value).strip() or "OTRO"

# =============================================================================
# UTILIDADES
# =============================================================================

def parse_ar_num(value) -> float:
    """Convierte números argentinos/mixtos a float."""
    if pd.isna(value):
        return 0.0
    s = str(value).strip()
    if s == "" or s.lower() in {"nan", "none", "null"}:
        return 0.0
    s = s.replace("$", "").replace(" ", "").replace("\u00a0", "")
    # Negativos tipo (1.234,56)
    neg = False
    if s.startswith("(") and s.endswith(")"):
        neg = True
        s = s[1:-1]
    # Si hay coma, asumimos formato AR: 1.234,56
    if "," in s:
        s = s.replace(".", "").replace(",", ".")
    try:
        out = float(s)
        return -out if neg else out
    except Exception:
        return 0.0


def norm_txt(x) -> str:
    """Normaliza texto para matcheo: mayúsculas, sin acentos y espacios simples."""
    if pd.isna(x):
        return ""
    s = str(x).strip().upper()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    return re.sub(r"\s+", " ", s)


def safe_date(x) -> Optional[pd.Timestamp]:
    if pd.isna(x):
        return pd.NaT
    if isinstance(x, (datetime, pd.Timestamp)):
        return pd.to_datetime(x, errors="coerce")
    s = str(x).strip()
    # En QR suele venir "dd/mm/yyyy hh:mm:ss"
    s10 = s[:10]
    for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
        try:
            return pd.to_datetime(datetime.strptime(s10, fmt))
        except Exception:
            pass
    return pd.to_datetime(s, dayfirst=True, errors="coerce")


def fmt_date(ts) -> str:
    ts = safe_date(ts)
    if pd.isna(ts):
        return ""
    return ts.strftime("%d/%m/%Y")


def amount_equal(a: float, b: float, tol: float = 1.00) -> bool:
    return abs(abs(float(a)) - abs(float(b))) <= tol


def classify_bank(concepto: str) -> str:
    c = norm_txt(concepto)
    if "CR DEBIN SPOT" in c:
        return "QR"
    if "CR LIQ" in c or ("AMEX" in c and "CR" in c):
        return "LIQUIDACION_TARJETA"
    if "DEB LIQ" in c or "DEB.LIQ" in c:
        return "DEBITO_LIQ_TARJETA"
    if "GRAVAMEN LEY 25413 S/DEB" in c:
        return "IMPUESTO_LEY25413_DEB"
    if "GRAVAMEN LEY 25413 S/CRED" in c:
        return "IMPUESTO_LEY25413_CRED"
    if "INGRESOS BRUTOS" in c or "IIBB" in c:
        return "RETENCION_IIBB"
    if "I.V.A. BASE" in c or "IVA BASE" in c:
        return "IVA"
    if "COM TRANSFE" in c or "COM. TRANSFE" in c:
        return "GASTO_BANCARIO"
    if "C BE TR O/BCO" in c or "TRANSF.INT.DIST.TITULAR" in c or "CRED.TRAN" in c:
        return "TRANSFERENCIA_ENTRANTE"
    if "DB CREDIN" in c or "DEB.TRAN.INTERBMISMO" in c or "DEBITO TRANSF" in c:
        return "TRANSFERENCIA_SALIENTE"
    if "DEBITO PAGO DIRECTO" in c or "PAGO DIRECTO" in c:
        return "DEBITO_PAGO_DIRECTO"
    return "OTRO"


def category_compatible(pending_cat: str, actual_cat: str, pending_side: str, actual_type: str = "") -> bool:
    """Compatibilidad flexible para regularizar pendientes anteriores."""
    if pending_cat == actual_cat:
        return True
    # Banco egreso anterior suele aparecer ahora como MB-EXT, sin categoría fina del lado Flexxus.
    if pending_side == "BANCO_NO_FLEXXUS" and actual_type in {"MB-EXT", "MB-ENT-EX"}:
        return True
    # Banco ingreso anterior puede aparecer como PAV.
    if pending_side == "BANCO_NO_FLEXXUS" and actual_type == "PAV":
        return True
    # Flexxus PAV anterior puede acreditarse como QR, tarjeta o transferencia.
    if pending_side == "FLEXXUS_NO_BANCO" and actual_cat in {"QR", "LIQUIDACION_TARJETA", "TRANSFERENCIA_ENTRANTE", "OTRO"}:
        return True
    # Flexxus egreso anterior puede debitarse como transferencia o débito.
    if pending_side == "FLEXXUS_NO_BANCO" and actual_cat in {"TRANSFERENCIA_SALIENTE", "DEBITO_PAGO_DIRECTO", "OTRO"}:
        return True
    return False

# =============================================================================
# PARSERS
# =============================================================================

def parse_flexxus(file) -> pd.DataFrame:
    df = pd.read_excel(file, dtype=str, header=None)
    rows = []
    for _, row in df.iterrows():
        fecha = str(row.iloc[0]).strip() if len(row) > 0 and pd.notna(row.iloc[0]) else ""
        if "/" not in fecha:
            continue
        tipo = str(row.iloc[1]).strip() if len(row) > 1 and pd.notna(row.iloc[1]) else ""
        numero = str(row.iloc[2]).strip() if len(row) > 2 and pd.notna(row.iloc[2]) else ""
        movimiento = str(row.iloc[4]).strip() if len(row) > 4 and pd.notna(row.iloc[4]) else ""
        debe = parse_ar_num(row.iloc[7]) if len(row) > 7 else 0.0
        haber = parse_ar_num(row.iloc[9]) if len(row) > 9 else 0.0
        saldo = parse_ar_num(row.iloc[10]) if len(row) > 10 else 0.0
        if tipo not in {"PAV", "MB-ENT-EX", "MB-EXT"}:
            continue
        # PAV suele ser Debe; egresos suelen estar en Haber. Tomamos el no-cero.
        monto = debe if abs(debe) > 0.005 else haber
        if abs(monto) <= 0.005:
            continue
        rows.append({
            "row_id": f"F-{uuid.uuid4().hex[:10]}",
            "FechaFlexxus": fecha,
            "FechaFlexxus_dt": safe_date(fecha),
            "Tipo": tipo,
            "Numero": numero,
            "Movimiento": movimiento,
            "ConceptoNorm": norm_txt(movimiento),
            "MontoFlexxus": abs(float(monto)),
            "SaldoFlexxus": saldo,
            "EsIngresoFlexxus": tipo == "PAV",
            "EsEgresoFlexxus": tipo in {"MB-ENT-EX", "MB-EXT"},
            "EsPedidosYa": "PEDIDOS YA" in norm_txt(movimiento),
            "Matched": False,
            "ConsumedPrev": False,
            "MatchStage": "",
            "MatchRef": "",
            "BancoFecha": "",
            "BancoConcepto": "",
            "BancoComprobante": "",
            "FechaAcreditacionUsada": "",
            "Diagnostico": "",
        })
    out = pd.DataFrame(rows)
    if out.empty:
        raise ValueError("No se encontraron movimientos PAV, MB-ENT-EX o MB-EXT en Flexxus.")
    return out


def parse_banco(file) -> pd.DataFrame:
    df = pd.read_excel(file, dtype=str, header=None)
    header_row = None
    for i, row in df.iterrows():
        if norm_txt(row.iloc[0] if len(row) > 0 else "") == "FECHA":
            header_row = i
            break
    if header_row is None:
        # fallback: usar fila 4 como en Banco Nación exportado
        header_row = 4
    raw = df.iloc[header_row + 1:].copy()
    raw = raw.iloc[:, :5]
    raw.columns = ["Fecha", "Comprobante", "Concepto", "Importe", "Saldo"]
    raw = raw.dropna(subset=["Fecha", "Concepto"])
    raw = raw[raw["Fecha"].astype(str).str.strip().ne("")]
    rows = []
    for source_order, (_, r) in enumerate(raw.iterrows()):
        imp = parse_ar_num(r["Importe"])
        if abs(imp) <= 0.005:
            continue
        rows.append({
            "row_id": f"B-{uuid.uuid4().hex[:10]}",
            "SourceOrder": source_order,
            "Fecha": str(r["Fecha"]).strip(),
            "Fecha_dt": safe_date(r["Fecha"]),
            "Comprobante": str(r["Comprobante"]).strip() if pd.notna(r["Comprobante"]) else "",
            "Concepto": str(r["Concepto"]).strip() if pd.notna(r["Concepto"]) else "",
            "ConceptoNorm": norm_txt(r["Concepto"]),
            "ImporteNum": imp,
            "ImporteAbs": abs(imp),
            "SaldoNum": parse_ar_num(r["Saldo"]),
            "EsIngreso": imp > 0,
            "EsEgreso": imp < 0,
            "Categoria": classify_bank(r["Concepto"]),
            "Matched": False,
            "ConsumedPrev": False,
            "MatchStage": "",
            "MatchRef": "",
            "FlexxusNumero": "",
            "Diagnostico": "",
        })
    out = pd.DataFrame(rows)
    if out.empty:
        raise ValueError("No se encontraron movimientos bancarios válidos.")
    return out.reset_index(drop=True)



def parse_qr(file) -> pd.DataFrame:
    if file is None:
        return pd.DataFrame()
    qr = pd.read_excel(file, dtype=str)
    if qr.empty:
        return qr

    def first_existing(cols):
        for c in cols:
            if c in qr.columns:
                return c
        normalized = {norm_txt(c): c for c in qr.columns}
        for c in cols:
            nc = norm_txt(c)
            if nc in normalized:
                return normalized[nc]
        return None

    def series_or_blank(col):
        if col and col in qr.columns:
            return qr[col].fillna("")
        return pd.Series([""] * len(qr), index=qr.index)

    monto_col = first_existing(["Monto total", "Monto Total", "MONTO TOTAL", "Monto", "Importe", "Importe total"])
    fecha_col = first_existing(["Fecha", "FECHA", "Fecha QR", "Fecha de pago", "Fecha transacción", "Fecha Transacción"])
    comercio_col = first_existing(["Cód. comercio", "Cod. comercio", "Código comercio", "Comercio", "COMERCIO"])
    ticket_col = first_existing(["Ticket", "TICKET", "Cupón", "Cupon", "Id QR", "ID QR"])

    qr["MontoTotal"] = series_or_blank(monto_col).apply(parse_ar_num) if monto_col else 0.0
    qr["NetoQR"] = qr["MontoTotal"] * 0.99032
    qr["FechaQR"] = series_or_blank(fecha_col).astype(str).str[:10] if fecha_col else ""
    qr["FechaQR_dt"] = qr["FechaQR"].apply(safe_date)
    qr["CodComercio"] = series_or_blank(comercio_col).astype(str) if comercio_col else ""
    qr["Cupon"] = series_or_blank(ticket_col).astype(str) if ticket_col else ""
    return qr

def parse_trx(file) -> pd.DataFrame:
    if file is None:
        return pd.DataFrame()
    trx = pd.read_excel(file, dtype=str)
    if trx.empty:
        return trx
    if "IMPORTE NETO" in trx.columns:
        trx["MontoNeto"] = trx["IMPORTE NETO"].apply(parse_ar_num)
    elif "TOTAL LIQUIDACION" in trx.columns:
        trx["MontoNeto"] = trx["TOTAL LIQUIDACION"].apply(parse_ar_num)
    else:
        trx["MontoNeto"] = 0.0
    if "FECHA DE PAGO" in trx.columns:
        trx["FechaPago_dt"] = trx["FECHA DE PAGO"].apply(safe_date)
    if "COMERCIO" in trx.columns:
        trx["Local"] = trx["COMERCIO"].astype(str).map(LOCAL_MAP).fillna(trx["COMERCIO"].astype(str))
    return trx

# =============================================================================
# CONCILIACIÓN ANTERIOR / PENDIENTES
# =============================================================================

def _find_header_row(df: pd.DataFrame, required_any: List[str]) -> Optional[int]:
    required = [norm_txt(r) for r in required_any]
    for i in range(min(len(df), 25)):
        row_txt = " | ".join(norm_txt(x) for x in df.iloc[i].tolist())
        if "FECHA FLEXXUS" in required:
            if "FECHA FLEXXUS" in row_txt and "MOVIMIENTO" in row_txt and ("MONTO" in row_txt or "IMPORTE" in row_txt):
                return i
            continue
        if "FECHA BANCO" in required:
            if "FECHA BANCO" in row_txt and "CONCEPTO BANCO" in row_txt and "IMPORTE" in row_txt:
                return i
            continue
        if "TIPO PENDIENTE" in required:
            if ("TIPO PENDIENTE" in row_txt or "ORIGEN" in row_txt) and ("MONTO" in row_txt or "IMPORTE" in row_txt):
                return i
            continue
        if all(r in row_txt for r in required):
            return i
    return None


def stable_pending_id(origen: str, fecha: str, tipo: str, numero: str, concepto: str, categoria: str, monto: float) -> str:
    """ID base determinístico para trazabilidad.

    V5.0: el ID final visible no queda solo en este hash. Luego
    assign_stable_pending_ids agrega una LETRA + secuencia de ocurrencia
    (-A001, -B002, etc.) para evitar duplicados incluso cuando Banco Nación
    trae dos líneas con misma fecha, comprobante, concepto e importe.
    """
    import hashlib
    cents = int(round(abs(float(monto or 0)) * 100))
    raw = "|".join([
        norm_txt(origen), norm_txt(fecha), norm_txt(tipo), norm_txt(numero),
        norm_txt(concepto), norm_txt(categoria), str(cents)
    ])
    return "PEND-" + hashlib.sha1(raw.encode("utf-8")).hexdigest()[:12].upper()

def _trace_letter(n: int) -> str:
    """Convierte 1->A, 2->B ... 26->Z, 27->AA para sufijos visibles."""
    n = int(n)
    letters = ""
    while n > 0:
        n, rem = divmod(n - 1, 26)
        letters = chr(65 + rem) + letters
    return letters or "A"

def make_trace_id(base_id: str, occurrence: int) -> str:
    """ID final único y legible: PEND-XXXXXXXXXXXX-A001."""
    occurrence = max(1, int(occurrence))
    return f"{base_id}-{_trace_letter(occurrence)}{occurrence:03d}"

def assign_stable_pending_ids(df: pd.DataFrame) -> pd.DataFrame:
    """Asigna IDs únicos visibles por movimiento.

    Conserva IDs V5 ya existentes si vienen de una conciliación anterior.
    Para registros nuevos calcula un ID base determinístico y agrega una letra
    + secuencia de ocurrencia. Esto evita colisiones en líneas bancarias
    idénticas y hace que cada pendiente sea individualmente trazable.
    """
    if df is None or df.empty:
        return df
    out = df.copy()

    # Si el archivo anterior ya trae IDs V5 únicos, se respetan.
    if "pending_id" in out.columns:
        existing = out["pending_id"].fillna("").astype(str)
        if existing.str.match(r"^PEND-[A-F0-9]{12}-[A-Z]+\d{3}$").all() and existing.is_unique:
            return out

    base_ids = []
    for _, r in out.iterrows():
        base_ids.append(stable_pending_id(
            r.get("Origen", ""), r.get("FechaOrigen", ""), r.get("TipoMovimiento", ""),
            r.get("Numero", ""), r.get("Concepto", ""), r.get("Categoria", ""), r.get("Monto", 0)
        ))

    seen = {}
    final_ids = []
    for base in base_ids:
        seen[base] = seen.get(base, 0) + 1
        final_ids.append(make_trace_id(base, seen[base]))

    # Defensa adicional: si por algún motivo quedara repetido, agrega sufijo X#### por posición.
    counts = {}
    unique_ids = []
    for i, pid in enumerate(final_ids, start=1):
        counts[pid] = counts.get(pid, 0) + 1
        if counts[pid] == 1:
            unique_ids.append(pid)
        else:
            unique_ids.append(f"{pid}-X{i:04d}")

    out["pending_id"] = unique_ids
    return out

def parse_previous_conciliation(file) -> Tuple[pd.DataFrame, Dict[str, float], str]:
    """Lee conciliación anterior nueva o vieja.

    Devuelve pendientes abiertos con signo lógico y resumen anterior.
    Si el archivo no tiene hoja 'Pendientes abiertos', intenta leer hojas del reporte viejo.
    """
    if file is None:
        return pd.DataFrame(columns=[
            "pending_id", "FechaOrigen", "Origen", "TipoPendiente", "TipoMovimiento",
            "Numero", "Concepto", "Categoria", "Monto", "SignoCalculo", "Estado", "Fuente"
        ]), {}, "APERTURA"

    xl = pd.ExcelFile(file)
    pendings = []
    prev = {}
    mode = "ANTERIOR"

    # Resumen anterior
    if "Conciliacion semanal" in xl.sheet_names:
        ws = pd.read_excel(file, sheet_name="Conciliacion semanal", header=None, dtype=str)
        for _, row in ws.iterrows():
            label = norm_txt(row.iloc[0] if len(row) > 0 else "")
            val = parse_ar_num(row.iloc[1] if len(row) > 1 else 0)
            if "SALDO SEGÚN EXTRACTO" in label or "SALDO S/EXTRACTO" in label:
                prev["saldo_banco_anterior"] = val
            elif "SALDO S/FLEXXUS" in label and "PASANDO" not in label and "CALCULADO" not in label:
                prev["saldo_flexxus_anterior"] = val
            elif "DIFERENCIA FINAL" in label:
                prev["diferencia_final_anterior"] = val

    # Formato nuevo: Pendientes abiertos
    if "Pendientes abiertos" in xl.sheet_names:
        pa_raw = pd.read_excel(file, sheet_name="Pendientes abiertos", header=None, dtype=str)
        h = _find_header_row(pa_raw, ["Tipo pendiente", "Origen", "Importe", "Monto"])
        if h is not None:
            pa = pd.read_excel(file, sheet_name="Pendientes abiertos", header=h, dtype=str)
            for _, r in pa.iterrows():
                concepto = str(r.get("Concepto", r.get("Movimiento", ""))).strip()
                monto = parse_ar_num(r.get("Importe", r.get("Monto", 0)))
                if not concepto or abs(monto) <= 0.005:
                    continue
                origen = str(r.get("Origen", r.get("Tipo pendiente", ""))).strip().upper()
                tipo_p = str(r.get("Tipo pendiente", origen)).strip().upper()
                categoria = str(r.get("Categoría", r.get("Categoria", "OTRO"))).strip() or "OTRO"
                signo = infer_pending_sign(origen, tipo_p, categoria, concepto, monto)
                pendings.append({
                    "pending_id": str(r.get("ID", "")) or f"P-{uuid.uuid4().hex[:10]}",
                    "FechaOrigen": str(r.get("Fecha origen", r.get("Fecha", ""))).strip(),
                    "Origen": normalize_pending_origin(origen, tipo_p),
                    "TipoPendiente": tipo_p,
                    "TipoMovimiento": str(r.get("Tipo", "")).strip(),
                    "Numero": str(r.get("Número", r.get("Numero", ""))).strip(),
                    "Concepto": concepto,
                    "Categoria": categoria,
                    "Monto": abs(monto),
                    "SignoCalculo": signo,
                    "Estado": "ABIERTO",
                    "Fuente": "Pendientes abiertos anterior",
                })

    # Formato viejo: reconstruir desde hojas detalle
    if not pendings:
        # Flexxus no Banco
        if "Flexxus no Banco" in xl.sheet_names:
            raw = pd.read_excel(file, sheet_name="Flexxus no Banco", header=None, dtype=str)
            h = _find_header_row(raw, ["Fecha Flexxus", "Movimiento", "Monto"])
            if h is not None:
                fnb = pd.read_excel(file, sheet_name="Flexxus no Banco", header=h, dtype=str)
                for _, r in fnb.iterrows():
                    concepto = str(r.get("Movimiento", "")).strip()
                    monto = parse_ar_num(r.get("Monto", 0))
                    tipo = str(r.get("Tipo", "")).strip()
                    if not concepto or abs(monto) <= 0.005:
                        continue
                    signo = -1 if tipo == "PAV" else 1
                    pendings.append({
                        "pending_id": f"P-{uuid.uuid4().hex[:10]}",
                        "FechaOrigen": str(r.get("Fecha Flexxus", "")).strip(),
                        "Origen": "FLEXXUS_NO_BANCO",
                        "TipoPendiente": "FLEXXUS_NO_BANCO",
                        "TipoMovimiento": tipo,
                        "Numero": str(r.get("Número", "")).strip(),
                        "Concepto": concepto,
                        "Categoria": "PAV" if tipo == "PAV" else "EGRESO_FLEXXUS",
                        "Monto": abs(monto),
                        "SignoCalculo": signo,
                        "Estado": "ABIERTO",
                        "Fuente": "Flexxus no Banco anterior",
                    })
        # Banco ingresos no Flexxus
        if "Banco ingresos no Flexxus" in xl.sheet_names:
            raw = pd.read_excel(file, sheet_name="Banco ingresos no Flexxus", header=None, dtype=str)
            h = _find_header_row(raw, ["Fecha banco", "Concepto banco", "Importe"])
            if h is not None:
                bi = pd.read_excel(file, sheet_name="Banco ingresos no Flexxus", header=h, dtype=str)
                for _, r in bi.iterrows():
                    concepto = str(r.get("Concepto banco", "")).strip()
                    monto = parse_ar_num(r.get("Importe", 0))
                    if not concepto or abs(monto) <= 0.005:
                        continue
                    categoria = str(r.get("Categoría", r.get("Categoria", ""))).strip() or classify_bank(concepto)
                    pendings.append({
                        "pending_id": f"P-{uuid.uuid4().hex[:10]}",
                        "FechaOrigen": str(r.get("Fecha banco", "")).strip(),
                        "Origen": "BANCO_NO_FLEXXUS",
                        "TipoPendiente": "BANCO_INGRESO_NO_FLEXXUS",
                        "TipoMovimiento": "PAV",
                        "Numero": str(r.get("Comprobante banco", "")).strip(),
                        "Concepto": concepto,
                        "Categoria": categoria,
                        "Monto": abs(monto),
                        "SignoCalculo": 1,
                        "Estado": "ABIERTO",
                        "Fuente": "Banco ingresos no Flexxus anterior",
                    })
        # Banco egresos no Flexxus
        if "Banco egresos no Flexxus" in xl.sheet_names:
            raw = pd.read_excel(file, sheet_name="Banco egresos no Flexxus", header=None, dtype=str)
            h = _find_header_row(raw, ["Fecha banco", "Concepto banco", "Importe"])
            if h is not None:
                be = pd.read_excel(file, sheet_name="Banco egresos no Flexxus", header=h, dtype=str)
                for _, r in be.iterrows():
                    concepto = str(r.get("Concepto banco", "")).strip()
                    monto = parse_ar_num(r.get("Importe", 0))
                    if not concepto or abs(monto) <= 0.005:
                        continue
                    categoria = classify_bank(concepto)
                    pendings.append({
                        "pending_id": f"P-{uuid.uuid4().hex[:10]}",
                        "FechaOrigen": str(r.get("Fecha banco", "")).strip(),
                        "Origen": "BANCO_NO_FLEXXUS",
                        "TipoPendiente": "BANCO_EGRESO_NO_FLEXXUS",
                        "TipoMovimiento": "MB-EXT" if categoria in IMP_CATS else "MB-ENT-EX",
                        "Numero": str(r.get("Comprobante banco", "")).strip(),
                        "Concepto": concepto,
                        "Categoria": categoria,
                        "Monto": abs(monto),
                        "SignoCalculo": -1,
                        "Estado": "ABIERTO",
                        "Fuente": "Banco egresos no Flexxus anterior",
                    })


    # Formato V4.8: movimientos ya procesados/importados en la conciliación anterior.
    # Sirve para reprocesos del mismo extracto: si el movimiento ya estuvo en "Carga Flexxus"
    # o en "Regularizaciones", no debe volver a entrar como C1/C2 corriente.
    if "Carga Flexxus" in xl.sheet_names:
        try:
            raw = pd.read_excel(file, sheet_name="Carga Flexxus", header=None, dtype=str)
            h = _find_header_row(raw, ["Fecha mov. Flexxus", "Nro. Flexxus", "Monto"])
            if h is not None:
                cf = pd.read_excel(file, sheet_name="Carga Flexxus", header=h, dtype=str)
                for _, r in cf.iterrows():
                    tipo = str(r.get("Tipo", "")).strip()
                    numero = str(r.get("Nro. Flexxus", r.get("Número", ""))).strip()
                    monto = parse_ar_num(r.get("Monto", 0))
                    fecha = str(r.get("Fecha mov. Flexxus", "")).strip()
                    if tipo in {"PAV", "MB-ENT-EX", "MB-EXT"} and numero and abs(monto) > 0.005:
                        pendings.append({
                            "pending_id": stable_pending_id("FLEXXUS_PROCESADO_ANTERIOR", fecha, tipo, numero, "", "PROCESADO_ANTERIOR", monto),
                            "FechaOrigen": fecha,
                            "Origen": "FLEXXUS_PROCESADO_ANTERIOR",
                            "TipoPendiente": "FLEXXUS_PROCESADO_ANTERIOR",
                            "TipoMovimiento": tipo,
                            "Numero": numero,
                            "Concepto": "Movimiento ya incluido en Carga Flexxus anterior",
                            "Categoria": "PROCESADO_ANTERIOR",
                            "Monto": abs(monto),
                            "SignoCalculo": 0,
                            "Estado": "PROCESADO_ANTERIOR",
                            "Fuente": "Carga Flexxus anterior",
                        })
                        banco_fecha = str(r.get("Fecha banco real", r.get("Fecha banco", ""))).strip()
                        banco_comp = str(r.get("Comprobante banco", "")).strip()
                        banco_conc = str(r.get("Concepto banco", "")).strip()
                        if banco_conc or banco_comp:
                            pendings.append({
                                "pending_id": stable_pending_id("BANCO_PROCESADO_ANTERIOR", banco_fecha, tipo, banco_comp, banco_conc, classify_bank(banco_conc), monto),
                                "FechaOrigen": banco_fecha,
                                "Origen": "BANCO_PROCESADO_ANTERIOR",
                                "TipoPendiente": "BANCO_PROCESADO_ANTERIOR",
                                "TipoMovimiento": tipo,
                                "Numero": banco_comp,
                                "Concepto": banco_conc,
                                "Categoria": classify_bank(banco_conc),
                                "Monto": abs(monto),
                                "SignoCalculo": 0,
                                "Estado": "PROCESADO_ANTERIOR",
                                "Fuente": "Banco usado en Carga Flexxus anterior",
                            })
        except Exception:
            pass

    if "Regularizaciones" in xl.sheet_names:
        try:
            raw = pd.read_excel(file, sheet_name="Regularizaciones", header=None, dtype=str)
            h = _find_header_row(raw, ["Fecha Flexxus", "Tipo", "Número", "Importe"])
            if h is not None:
                rg = pd.read_excel(file, sheet_name="Regularizaciones", header=h, dtype=str)
                for _, r in rg.iterrows():
                    tipo = str(r.get("Tipo", "")).strip()
                    numero = str(r.get("Número", r.get("Numero", ""))).strip()
                    mov = str(r.get("Movimiento", "")).strip()
                    monto = parse_ar_num(r.get("Importe", r.get("Monto", 0)))
                    fecha = str(r.get("Fecha Flexxus", r.get("Fecha", ""))).strip()
                    if tipo in {"PAV", "MB-ENT-EX", "MB-EXT"} and numero and abs(monto) > 0.005:
                        pendings.append({
                            "pending_id": stable_pending_id("FLEXXUS_PROCESADO_ANTERIOR", fecha, tipo, numero, mov, "PROCESADO_ANTERIOR", monto),
                            "FechaOrigen": fecha,
                            "Origen": "FLEXXUS_PROCESADO_ANTERIOR",
                            "TipoPendiente": "FLEXXUS_PROCESADO_ANTERIOR",
                            "TipoMovimiento": tipo,
                            "Numero": numero,
                            "Concepto": mov or "Movimiento ya incluido en Regularizaciones anteriores",
                            "Categoria": "PROCESADO_ANTERIOR",
                            "Monto": abs(monto),
                            "SignoCalculo": 0,
                            "Estado": "PROCESADO_ANTERIOR",
                            "Fuente": "Regularizaciones anteriores",
                        })
        except Exception:
            pass



    # Si la conciliación anterior cerró con diferencia distinta de cero, no se descarta:
    # se arrastra explícitamente como pendiente técnico anterior con ID estable.
    # Esto evita perder trazabilidad cuando una versión anterior cerró mal (ej. V4.7 tercer registro).
    try:
        prev_diff = float(prev.get("diferencia_final_anterior", 0) or 0)
        if abs(prev_diff) > 1.0:
            pendings.append({
                "pending_id": stable_pending_id("DIFERENCIA_FINAL_ANTERIOR", "", "", "DIF_ANT", "Diferencia final anterior abierta", "DIFERENCIA_FINAL_ANTERIOR", prev_diff),
                "FechaOrigen": "",
                "Origen": "DIFERENCIA_FINAL_ANTERIOR",
                "TipoPendiente": "DIFERENCIA_FINAL_ANTERIOR",
                "TipoMovimiento": "",
                "Numero": "DIF_ANT",
                "Concepto": "Diferencia final de la conciliación anterior aún abierta",
                "Categoria": "DIFERENCIA_FINAL_ANTERIOR",
                "Monto": abs(prev_diff),
                "SignoCalculo": 1 if prev_diff > 0 else -1,
                "Estado": "ABIERTO_ANTERIOR",
                "Fuente": "Diferencia final anterior explícita",
            })
    except Exception:
        pass


    out = pd.DataFrame(pendings)
    if out.empty:
        out = pd.DataFrame(columns=[
            "pending_id", "FechaOrigen", "Origen", "TipoPendiente", "TipoMovimiento",
            "Numero", "Concepto", "Categoria", "Monto", "SignoCalculo", "Estado", "Fuente"
        ])
    out = assign_stable_pending_ids(out)
    return out, prev, mode


def normalize_pending_origin(origen: str, tipo_p: str) -> str:
    txt = norm_txt(origen + " " + tipo_p)
    if "FLEXXUS" in txt and "BANCO" in txt:
        return "FLEXXUS_NO_BANCO" if txt.index("FLEXXUS") < txt.index("BANCO") else "BANCO_NO_FLEXXUS"
    if "BANCO" in txt and "FLEXXUS" in txt:
        return "BANCO_NO_FLEXXUS"
    if "FLEXXUS" in txt:
        return "FLEXXUS_NO_BANCO"
    return "BANCO_NO_FLEXXUS"


def infer_pending_sign(origen: str, tipo_p: str, categoria: str, concepto: str, monto: float) -> int:
    txt = norm_txt(origen + " " + tipo_p + " " + categoria + " " + concepto)
    if "BANCO" in txt and "INGRES" in txt:
        return 1
    if "BANCO" in txt and ("EGRES" in txt or "DEBIT" in txt or "RETENC" in txt or "IMPUEST" in txt or "IVA" in txt):
        return -1
    if "FLEXXUS" in txt and "INGRES" in txt:
        return -1
    if "FLEXXUS" in txt and "EGRES" in txt:
        return 1
    if categoria in {"QR", "LIQUIDACION_TARJETA", "TRANSFERENCIA_ENTRANTE"}:
        return 1
    if categoria in IMP_CATS or categoria in {"TRANSFERENCIA_SALIENTE"}:
        return -1
    return 1 if monto >= 0 else -1


def classify_mbext(movimiento: str) -> str:
    """Clasifica una fila MB-EXT de Flexxus para matcheo agregado contra banco.

    V4.4: normaliza acentos y puntos para reconocer textos como
    "Ley 25.413 déb" y "Ley 25.413 cré".
    """
    c = norm_txt(movimiento)
    compact = re.sub(r"[^A-Z0-9]", "", c)
    if "INGRESOSBRUTOS" in compact or "IIBB" in compact or "IB" == compact[:2] or "RETENCIONIIBB" in compact:
        return "RETENCION_IIBB"
    if "25413" in compact:
        if "DEB" in compact or "DEBIT" in compact:
            return "IMPUESTO_LEY25413_DEB"
        if "CRE" in compact or "CRED" in compact or "CREDIT" in compact:
            return "IMPUESTO_LEY25413_CRED"
    if "IVA" in compact:
        return "IVA"
    if "COMTRANSFE" in compact or "COMISION" in compact or "COMBANC" in compact:
        return "GASTO_BANCARIO"
    if "DEBLIQ" in compact or "LIQMASTER" in compact or "MASTERCARD" in compact:
        return "DEBITO_LIQ_TARJETA"
    if "PAGODIRECTO" in compact:
        return "DEBITO_PAGO_DIRECTO"
    return "OTRO_MBEXT"

def amount_close(a: float, b: float, base_tol: float = 2.0, rel_tol: float = 0.00002) -> bool:
    """Tolerancia mixta: mínimo fijo + margen proporcional para grandes montos."""
    a = abs(float(a)); b = abs(float(b))
    return abs(a - b) <= max(base_tol, max(a, b) * rel_tol)


def amount_exact(a: float, b: float, tol: float = 0.05) -> bool:
    """Match 1:1 estricto. Evita compensar QR/tarjetas con diferencias de centavos o importes casi iguales."""
    a = abs(float(a)); b = abs(float(b))
    return abs(a - b) <= tol


def get_saldo_banco_final(bank: pd.DataFrame) -> float:
    """Obtiene saldo final del extracto sin depender de bank.iloc[0]."""
    if bank.empty:
        return 0.0
    ordered = bank.sort_values("SourceOrder") if "SourceOrder" in bank.columns else bank.copy()
    first_date = ordered.iloc[0]["Fecha_dt"]
    last_date = ordered.iloc[-1]["Fecha_dt"]
    if pd.notna(first_date) and pd.notna(last_date) and first_date >= last_date:
        return float(ordered.iloc[0]["SaldoNum"])
    return float(ordered.iloc[-1]["SaldoNum"])


def get_saldo_banco_apertura(bank: pd.DataFrame) -> float:
    """Saldo de apertura del extracto actual.

    En Banco Nación el extracto suele venir en orden descendente y, cuando todos
    los movimientos son del mismo día, SourceOrder no permite reconstruir la
    primera operación cronológica. Por eso la apertura más robusta para control
    de continuidad es:
        saldo final del extracto - suma neta de movimientos del extracto.
    """
    if bank.empty:
        return 0.0
    try:
        return round(float(get_saldo_banco_final(bank)) - float(bank["ImporteNum"].sum()), 2)
    except Exception:
        tmp = bank.copy()
        tmp = tmp[pd.notna(tmp["Fecha_dt"])]
        if tmp.empty:
            return 0.0
        sort_cols = ["Fecha_dt"] + (["SourceOrder"] if "SourceOrder" in tmp.columns else ["row_id"])
        first = tmp.sort_values(sort_cols).iloc[0]
        return round(float(first["SaldoNum"]) - float(first["ImporteNum"]), 2)

def build_continuity_control(prev_summary: Dict[str, float], bank: pd.DataFrame) -> Dict:
    """Controla continuidad banco vs banco, no banco anterior vs Flexxus actual."""
    prev_bank = float(prev_summary.get("saldo_banco_anterior", 0) or 0)
    opening = get_saldo_banco_apertura(bank)
    if not prev_bank:
        return {"aplica": False, "mensaje": "Sin saldo banco anterior leído para validar continuidad."}
    diff = round(opening - prev_bank, 2)
    ok = abs(diff) <= 10.0
    return {
        "aplica": True,
        "saldo_banco_cierre_anterior": prev_bank,
        "saldo_banco_apertura_actual": opening,
        "diferencia_continuidad_banco": diff,
        "ok": ok,
        "mensaje": (
            "OK: apertura bancaria actual coincide con cierre anterior dentro de tolerancia operativa."
            if ok else
            "Revisar: la apertura bancaria actual no coincide con el cierre anterior. Puede faltar un movimiento o el rango del extracto no es consecutivo."
        ),
    }

# =============================================================================
# MOTOR DE MATCHING
# =============================================================================

def match_previous_pendings(prev_open: pd.DataFrame, flex: pd.DataFrame, bank: pd.DataFrame, prev_summary: Optional[Dict[str, float]] = None) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """Cancela pendientes anteriores contra movimientos actuales.

    V3 agrega matching agregado: una MB-EXT consolidada puede regularizar varias
    líneas pendientes anteriores de Banco no Flexxus de la misma categoría.
    """
    regs = []
    flex.attrs["has_previous_conciliation"] = not prev_open.empty
    if prev_open.empty:
        return prev_open.copy(), flex, bank, pd.DataFrame(regs)

    prev = prev_open.copy().reset_index(drop=True)
    prev_bank_final = float((prev_summary or {}).get("saldo_banco_anterior", 0) or 0)
    current_bank_final = get_saldo_banco_final(bank)
    is_reprocess_same_extract = bool(prev_bank_final and abs(prev_bank_final - current_bank_final) <= 10.0)
    flex.attrs["is_reprocess_same_extract"] = is_reprocess_same_extract
    bank.attrs["is_reprocess_same_extract"] = is_reprocess_same_extract

    processed_prev = prev[prev.get("Origen", "").astype(str).eq("FLEXXUS_PROCESADO_ANTERIOR")].copy() if ("Origen" in prev.columns and is_reprocess_same_extract) else pd.DataFrame()
    if not processed_prev.empty:
        for _, p in processed_prev.iterrows():
            tipo = str(p.get("TipoMovimiento", ""))
            numero = str(p.get("Numero", ""))
            monto = float(p.get("Monto", 0.0) or 0.0)
            candidates = flex[
                (~flex["Matched"]) & (~flex["ConsumedPrev"]) &
                (flex["Tipo"].astype(str).eq(tipo)) &
                (flex["Numero"].astype(str).eq(numero)) &
                (flex["MontoFlexxus"].apply(lambda x: amount_exact(x, monto, tol=0.05)))
            ].copy()
            if not candidates.empty:
                fidx = candidates.index[0]
                flex.loc[fidx, "ConsumedPrev"] = True
                flex.loc[fidx, "MatchStage"] = "YA_PROCESADO_ANTERIOR"
                flex.loc[fidx, "MatchRef"] = p.get("pending_id", "")
                flex.loc[fidx, "Diagnostico"] = "Movimiento ya incluido en conciliación anterior; no entra como pendiente corriente"
        prev = prev[~prev.get("Origen", "").astype(str).eq("FLEXXUS_PROCESADO_ANTERIOR")].reset_index(drop=True)

    processed_bank_prev = prev[prev.get("Origen", "").astype(str).eq("BANCO_PROCESADO_ANTERIOR")].copy() if ("Origen" in prev.columns and is_reprocess_same_extract) else pd.DataFrame()
    if not processed_bank_prev.empty:
        for _, p in processed_bank_prev.iterrows():
            monto = float(p.get("Monto", 0.0) or 0.0)
            comp = str(p.get("Numero", ""))
            cat = str(p.get("Categoria", ""))
            candidates = bank[
                (~bank["Matched"]) & (~bank["ConsumedPrev"]) &
                (bank["ImporteAbs"].apply(lambda x: amount_exact(x, monto, tol=0.05)))
            ].copy()
            if comp:
                exact_comp = candidates[candidates["Comprobante"].astype(str).eq(comp)]
                if not exact_comp.empty:
                    candidates = exact_comp
            if cat:
                same_cat = candidates[candidates["Categoria"].astype(str).eq(cat)]
                if not same_cat.empty:
                    candidates = same_cat
            if not candidates.empty:
                bidx = candidates.index[0]
                bank.loc[bidx, "ConsumedPrev"] = True
                bank.loc[bidx, "MatchStage"] = "BANCO_YA_PROCESADO_ANTERIOR"
                bank.loc[bidx, "MatchRef"] = p.get("pending_id", "")
                bank.loc[bidx, "Diagnostico"] = "Movimiento bancario ya usado en conciliación anterior; no entra como C3/C4 corriente"
        prev = prev[~prev.get("Origen", "").astype(str).eq("BANCO_PROCESADO_ANTERIOR")].reset_index(drop=True)

    prev["Regularizado"] = False
    prev["RegularizadoCon"] = ""
    prev["FechaRegularizacion"] = ""
    prev["Diagnostico"] = "Pendiente anterior abierto"

    # V4.9: regulariza Pedidos Ya pendiente anterior contra transferencia entrante actual.
    # Se hace por concepto + categoría bancaria + tolerancia chica, no por QR/tarjeta.
    pedidos_prev_idx = prev[
        (~prev["Regularizado"]) &
        (prev["Origen"].astype(str).eq("FLEXXUS_NO_BANCO")) &
        (prev["SignoCalculo"].astype(float) < 0) &
        (prev["Concepto"].astype(str).apply(lambda x: "PEDIDOS YA" in norm_txt(x)))
    ].index.tolist()
    for pidx in pedidos_prev_idx:
        monto = float(prev.loc[pidx, "Monto"])
        candidates = bank[
            (~bank["Matched"]) & (~bank["ConsumedPrev"]) &
            (bank["EsIngreso"]) &
            (bank["Categoria"].astype(str).eq("TRANSFERENCIA_ENTRANTE")) &
            (bank["ImporteAbs"].apply(lambda x: abs(float(x) - monto) <= 1.00))
        ].copy()
        if candidates.empty:
            continue
        candidates["date_diff"] = (candidates["Fecha_dt"] - safe_date(prev.loc[pidx, "FechaOrigen"])).abs()
        bidx = candidates.sort_values(["date_diff", "row_id"]).index[0]
        diff_py = round(float(bank.loc[bidx, "ImporteAbs"]) - monto, 2)
        bank.loc[bidx, "ConsumedPrev"] = True
        bank.loc[bidx, "MatchStage"] = "REGULARIZACION_ANTERIOR_PEDIDOSYA"
        bank.loc[bidx, "MatchRef"] = prev.loc[pidx, "pending_id"]
        bank.loc[bidx, "Diagnostico"] = f"Regulariza Pedidos Ya anterior; diferencia redondeo {diff_py:.2f}. No QR/no tarjeta"
        prev.loc[pidx, "Regularizado"] = True
        prev.loc[pidx, "RegularizadoCon"] = bank.loc[bidx, "row_id"]
        prev.loc[pidx, "FechaRegularizacion"] = bank.loc[bidx, "Fecha"]
        prev.loc[pidx, "Diagnostico"] = "Regularizado con transferencia entrante Pedidos Ya actual"
        regs.append({
            "ID pendiente": prev.loc[pidx, "pending_id"],
            "Fecha origen": prev.loc[pidx, "FechaOrigen"],
            "Origen pendiente": prev.loc[pidx, "TipoPendiente"],
            "Concepto pendiente": prev.loc[pidx, "Concepto"],
            "Categoría pendiente": prev.loc[pidx, "Categoria"],
            "Monto pendiente": monto,
            "Regularizado con": "Banco",
            "Fecha regularización": bank.loc[bidx, "Fecha"],
            "Tipo actual": "TRANSFERENCIA_ENTRANTE",
            "Referencia actual": bank.loc[bidx, "Comprobante"],
            "Concepto actual": bank.loc[bidx, "Concepto"],
            "Monto actual": float(bank.loc[bidx, "ImporteAbs"]),
            "Diferencia": diff_py,
            "Diagnóstico": "Regularización Pedidos Ya anterior contra transferencia entrante; sin QR ni tarjeta",
        })

    # 1) Matching agregado de MB-EXT actual contra múltiples pendientes bancarios anteriores.
    mbext_idx = flex[(~flex["Matched"]) & (~flex["ConsumedPrev"]) & (flex["Tipo"] == "MB-EXT")].index.tolist()
    for fidx in mbext_idx:
        monto = float(flex.loc[fidx, "MontoFlexxus"])
        cat = classify_mbext(flex.loc[fidx, "Movimiento"])
        if cat == "OTRO_MBEXT":
            continue
        candidates = prev[
            (~prev["Regularizado"]) &
            (prev["Origen"] == "BANCO_NO_FLEXXUS") &
            (prev["SignoCalculo"].astype(float) < 0) &
            (prev["Categoria"].astype(str).eq(cat))
        ].copy()
        if candidates.empty:
            continue
        total = float(candidates["Monto"].sum())
        if amount_close(total, monto):
            flex.loc[fidx, "ConsumedPrev"] = True
            flex.loc[fidx, "MatchStage"] = "REGULARIZACION_ANTERIOR_AGREGADA"
            flex.loc[fidx, "MatchRef"] = ",".join(candidates["pending_id"].astype(str).tolist())
            flex.loc[fidx, "Diagnostico"] = f"Regulariza {len(candidates)} pendientes bancarios anteriores categoría {cat} por suma agregada"
            flex.loc[fidx, "FechaAcreditacionUsada"] = flex.loc[fidx, "FechaFlexxus"]
            prev.loc[candidates.index, "Regularizado"] = True
            prev.loc[candidates.index, "RegularizadoCon"] = flex.loc[fidx, "row_id"]
            prev.loc[candidates.index, "FechaRegularizacion"] = flex.loc[fidx, "FechaFlexxus"]
            prev.loc[candidates.index, "Diagnostico"] = f"Regularizado por MB-EXT consolidada {flex.loc[fidx, 'Numero']}"
            regs.append({
                "ID pendiente": ",".join(candidates["pending_id"].astype(str).tolist()),
                "Fecha origen": "Varias",
                "Origen pendiente": "BANCO_EGRESO_NO_FLEXXUS_AGREGADO",
                "Concepto pendiente": f"{len(candidates)} líneas pendientes anteriores categoría {cat}",
                "Categoría pendiente": cat,
                "Monto pendiente": total,
                "Regularizado con": "Flexxus",
                "Fecha regularización": flex.loc[fidx, "FechaFlexxus"],
                "Tipo actual": "MB-EXT",
                "Referencia actual": flex.loc[fidx, "Numero"],
                "Concepto actual": flex.loc[fidx, "Movimiento"],
                "Monto actual": monto,
                "Diferencia": round(monto - total, 2),
                "Diagnóstico": f"Regularización agregada MB-EXT contra {len(candidates)} pendientes anteriores",
            })

    # 2) Matching 1:1 normal para lo que no fue agregado.
    prev = prev.sort_values(["Monto", "Categoria"], ascending=[False, True]).reset_index(drop=True)

    for pidx, p in prev.iterrows():
        if bool(p.get("Regularizado", False)):
            continue
        monto = float(p["Monto"])
        signo = int(p.get("SignoCalculo", 0))
        origen = str(p.get("Origen", ""))
        cat = str(p.get("Categoria", "OTRO"))

        if origen == "BANCO_NO_FLEXXUS":
            if signo > 0:
                candidates = flex[(~flex["Matched"]) & (~flex["ConsumedPrev"]) & (flex["Tipo"] == "PAV")].copy()
            else:
                candidates = flex[(~flex["Matched"]) & (~flex["ConsumedPrev"]) & (flex["Tipo"].isin(["MB-EXT", "MB-ENT-EX"]))].copy()
            # V4.4: si el pendiente anterior era un ingreso bancario que ahora
            # aparece como PAV, el match debe ser estricto. Evita cancelar QR
            # anteriores contra PAV actuales con diferencias chicas (ej. 0,24).
            if signo > 0:
                candidates = candidates[candidates["MontoFlexxus"].apply(lambda x: amount_exact(x, monto, tol=0.05))]
            else:
                candidates = candidates[candidates["MontoFlexxus"].apply(lambda x: amount_close(x, monto))]
            if not candidates.empty:
                candidates["date_diff"] = (candidates["FechaFlexxus_dt"] - safe_date(p.get("FechaOrigen", ""))).abs()
                best_idx = candidates.sort_values(["date_diff", "row_id"]).index[0]
                flex.loc[best_idx, "ConsumedPrev"] = True
                flex.loc[best_idx, "MatchStage"] = "REGULARIZACION_ANTERIOR"
                flex.loc[best_idx, "MatchRef"] = p["pending_id"]
                flex.loc[best_idx, "Diagnostico"] = "Regulariza Banco no Flexxus anterior"
                flex.loc[best_idx, "FechaAcreditacionUsada"] = flex.loc[best_idx, "FechaFlexxus"]
                prev.loc[pidx, "Regularizado"] = True
                prev.loc[pidx, "RegularizadoCon"] = flex.loc[best_idx, "row_id"]
                prev.loc[pidx, "FechaRegularizacion"] = flex.loc[best_idx, "FechaFlexxus"]
                prev.loc[pidx, "Diagnostico"] = "Regularizado con Flexxus actual"
                regs.append(build_reg_row(p, "Flexxus", flex.loc[best_idx].to_dict(), "Regulariza Banco no Flexxus anterior"))
                continue

        if origen == "FLEXXUS_NO_BANCO":
            if signo < 0:
                candidates = bank[(~bank["Matched"]) & (~bank["ConsumedPrev"]) & (bank["EsIngreso"])].copy()
            else:
                candidates = bank[(~bank["Matched"]) & (~bank["ConsumedPrev"]) & (bank["EsEgreso"])].copy()
            candidates = candidates[candidates["ImporteAbs"].apply(lambda x: amount_exact(x, monto))]
            if not candidates.empty:
                comp = candidates[candidates["Categoria"].apply(lambda c: category_compatible(cat, c, origen))]
                if comp.empty:
                    comp = candidates
                comp["date_diff"] = (comp["Fecha_dt"] - safe_date(p.get("FechaOrigen", ""))).abs()
                best_idx = comp.sort_values(["date_diff", "row_id"]).index[0]
                bank.loc[best_idx, "ConsumedPrev"] = True
                bank.loc[best_idx, "MatchStage"] = "REGULARIZACION_ANTERIOR"
                bank.loc[best_idx, "MatchRef"] = p["pending_id"]
                bank.loc[best_idx, "Diagnostico"] = "Regulariza Flexxus no Banco anterior"
                prev.loc[pidx, "Regularizado"] = True
                prev.loc[pidx, "RegularizadoCon"] = bank.loc[best_idx, "row_id"]
                prev.loc[pidx, "FechaRegularizacion"] = bank.loc[best_idx, "Fecha"]
                prev.loc[pidx, "Diagnostico"] = "Regularizado con Banco actual"
                regs.append(build_reg_row(p, "Banco", bank.loc[best_idx].to_dict(), "Regulariza Flexxus no Banco anterior"))
                continue

    # 3) Si un pendiente anterior sigue apareciendo igual en el archivo actual,
    # no es nuevo C1/C3/C4: se marca como arrastre del mismo pendiente.
    for pidx, p in prev.iterrows():
        if bool(p.get("Regularizado", False)):
            continue
        monto = float(p.get("Monto", 0.0) or 0.0)
        origen = str(p.get("Origen", ""))
        fecha = str(p.get("FechaOrigen", ""))
        numero = str(p.get("Numero", ""))
        concepto = norm_txt(p.get("Concepto", ""))
        if origen == "FLEXXUS_NO_BANCO":
            cand = flex[(~flex["Matched"]) & (~flex["ConsumedPrev"]) & flex["MontoFlexxus"].apply(lambda x: amount_exact(x, monto, tol=0.05))].copy()
            if numero:
                cand_num = cand[cand["Numero"].astype(str).eq(numero)]
                if not cand_num.empty:
                    cand = cand_num
            if not cand.empty:
                cand["same_date"] = cand["FechaFlexxus"].astype(str).eq(fecha).astype(int)
                best_idx = cand.sort_values(["same_date", "row_id"], ascending=[False, True]).index[0]
                flex.loc[best_idx, "ConsumedPrev"] = True
                flex.loc[best_idx, "MatchStage"] = "PENDIENTE_ANTERIOR_REPETIDO"
                flex.loc[best_idx, "MatchRef"] = p.get("pending_id", "")
                flex.loc[best_idx, "Diagnostico"] = "Sigue abierto desde conciliación anterior; no se cuenta como C1 nuevo"
        elif origen == "BANCO_NO_FLEXXUS":
            cand = bank[(~bank["Matched"]) & (~bank["ConsumedPrev"]) & bank["ImporteAbs"].apply(lambda x: amount_exact(x, monto, tol=0.05))].copy()
            if numero:
                cand_num = cand[cand["Comprobante"].astype(str).eq(numero)]
                if not cand_num.empty:
                    cand = cand_num
            if not cand.empty:
                cand["same_date"] = cand["Fecha"].astype(str).eq(fecha).astype(int)
                cand["same_concept"] = cand["ConceptoNorm"].astype(str).eq(concepto).astype(int)
                best_idx = cand.sort_values(["same_date", "same_concept", "row_id"], ascending=[False, False, True]).index[0]
                bank.loc[best_idx, "ConsumedPrev"] = True
                bank.loc[best_idx, "MatchStage"] = "PENDIENTE_ANTERIOR_REPETIDO"
                bank.loc[best_idx, "MatchRef"] = p.get("pending_id", "")
                bank.loc[best_idx, "Diagnostico"] = "Sigue abierto desde conciliación anterior; no se cuenta como C3/C4 nuevo"

    regs_df = pd.DataFrame(regs)
    return prev, flex, bank, regs_df

def build_reg_row(p, actual_source: str, actual: dict, diag: str) -> dict:
    if actual_source == "Flexxus":
        actual_fecha = actual.get("FechaFlexxus", "")
        actual_tipo = actual.get("Tipo", "")
        actual_concepto = actual.get("Movimiento", "")
        actual_monto = actual.get("MontoFlexxus", 0.0)
        actual_ref = actual.get("Numero", "")
    else:
        actual_fecha = actual.get("Fecha", "")
        actual_tipo = actual.get("Categoria", "")
        actual_concepto = actual.get("Concepto", "")
        actual_monto = actual.get("ImporteAbs", 0.0)
        actual_ref = actual.get("Comprobante", "")
    return {
        "ID pendiente": p.get("pending_id", ""),
        "Fecha origen": p.get("FechaOrigen", ""),
        "Origen pendiente": p.get("TipoPendiente", p.get("Origen", "")),
        "Concepto pendiente": p.get("Concepto", ""),
        "Categoría pendiente": p.get("Categoria", ""),
        "Monto pendiente": float(p.get("Monto", 0.0)),
        "Regularizado con": actual_source,
        "Fecha regularización": actual_fecha,
        "Tipo actual": actual_tipo,
        "Referencia actual": actual_ref,
        "Concepto actual": actual_concepto,
        "Monto actual": float(actual_monto),
        "Diferencia": round(float(actual_monto) - float(p.get("Monto", 0.0)), 2),
        "Diagnóstico": diag,
    }


def match_mbext_current_aggregated(flex: pd.DataFrame, bank: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """Matchea MB-EXT corrientes consolidadas contra N líneas bancarias de la misma categoría.

    Esto resuelve el caso real Dancona: Flexxus consolida semanalmente IIBB/IVA/Ley 25413,
    pero Banco Nación lo muestra en muchas líneas diarias.
    """
    rows = []
    for fidx in flex[(~flex["Matched"]) & (~flex["ConsumedPrev"]) & (flex["Tipo"] == "MB-EXT")].index.tolist():
        monto = float(flex.loc[fidx, "MontoFlexxus"])
        cat = classify_mbext(flex.loc[fidx, "Movimiento"])
        if cat == "OTRO_MBEXT":
            continue
        candidates = bank[
            (~bank["Matched"]) & (~bank["ConsumedPrev"]) &
            (bank["EsEgreso"]) & (bank["Categoria"] == cat)
        ].copy()
        if candidates.empty:
            continue

        fecha_f = flex.loc[fidx, "FechaFlexxus_dt"]
        if pd.notna(fecha_f):
            win = candidates[(candidates["Fecha_dt"] <= fecha_f + pd.Timedelta(days=1)) & (candidates["Fecha_dt"] >= fecha_f - pd.Timedelta(days=21))]
        else:
            win = candidates
        if win.empty:
            win = candidates

        total_win = float(win["ImporteAbs"].sum())
        used = win
        if not amount_close(total_win, monto):
            total_all = float(candidates["ImporteAbs"].sum())
            if amount_close(total_all, monto):
                used = candidates
                total_win = total_all
            else:
                continue

        flex.loc[fidx, "Matched"] = True
        flex.loc[fidx, "MatchStage"] = "CORRIENTE_AGREGADO_MBEXT"
        flex.loc[fidx, "MatchRef"] = ",".join(used["row_id"].astype(str).tolist())
        flex.loc[fidx, "BancoFecha"] = fmt_date(used["Fecha_dt"].max())
        flex.loc[fidx, "BancoConcepto"] = f"Consolidado banco {cat}: {len(used)} líneas"
        flex.loc[fidx, "BancoComprobante"] = "AGREGADO"
        flex.loc[fidx, "FechaAcreditacionUsada"] = flex.loc[fidx, "FechaFlexxus"]
        flex.loc[fidx, "Diagnostico"] = f"OK - MB-EXT agregado contra {len(used)} líneas bancarias categoría {cat}"

        bank.loc[used.index, "Matched"] = True
        bank.loc[used.index, "MatchStage"] = "CORRIENTE_AGREGADO_MBEXT"
        bank.loc[used.index, "MatchRef"] = flex.loc[fidx, "row_id"]
        bank.loc[used.index, "FlexxusNumero"] = flex.loc[fidx, "Numero"]
        bank.loc[used.index, "Diagnostico"] = f"Incluido en MB-EXT agregado {flex.loc[fidx, 'Numero']}"
        rows.append({
            "Fecha Flexxus": flex.loc[fidx, "FechaFlexxus"],
            "Número Flexxus": flex.loc[fidx, "Numero"],
            "Categoría": cat,
            "Concepto Flexxus": flex.loc[fidx, "Movimiento"],
            "Monto Flexxus": monto,
            "Cantidad líneas banco": len(used),
            "Total banco": total_win,
            "Diferencia": round(monto - total_win, 2),
            "Diagnóstico": "OK agregado MB-EXT",
        })
    return flex, bank, pd.DataFrame(rows)


def match_pav_qr_current_aggregated(flex: pd.DataFrame, bank: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """Matchea un PAV consolidado contra N QR bancarios del mismo día.

    Caso Dancona: Flexxus puede traer un PAV grande de regularización QR del período
    anterior, mientras Banco Nación muestra muchos CR DEBIN SPOT individuales.
    Se ejecuta después del match 1:1 estricto para que solo sume QR remanentes.
    """
    rows = []
    flex_candidates = flex[
        (~flex["Matched"]) & (~flex["ConsumedPrev"]) &
        (flex["Tipo"] == "PAV") &
        (~flex["EsPedidosYa"])
    ].copy()
    if flex_candidates.empty:
        return flex, bank, pd.DataFrame(rows)

    bank_qr = bank[
        (~bank["Matched"]) & (~bank["ConsumedPrev"]) &
        (bank["EsIngreso"]) & (bank["Categoria"] == "QR")
    ].copy()
    if bank_qr.empty:
        return flex, bank, pd.DataFrame(rows)

    for fecha_b, grp in bank_qr.groupby("Fecha"):
        if grp.empty or len(grp) < 2:
            continue
        total = float(grp["ImporteAbs"].sum())
        if total <= 0:
            continue
        same_day = flex_candidates[
            (~flex_candidates.index.isin([r.get("flex_index") for r in rows if "flex_index" in r])) &
            (flex_candidates["FechaFlexxus"] == fecha_b) &
            (flex_candidates["MontoFlexxus"].apply(lambda x: amount_close(x, total, base_tol=5.0, rel_tol=0.00005)))
        ].copy()
        if same_day.empty:
            continue
        same_day["monto_diff"] = same_day["MontoFlexxus"].apply(lambda x: abs(float(x) - total))
        fidx = same_day.sort_values(["monto_diff", "row_id"]).index[0]
        monto = float(flex.loc[fidx, "MontoFlexxus"])

        flex.loc[fidx, "Matched"] = True
        flex.loc[fidx, "MatchStage"] = "CORRIENTE_AGREGADO_QR"
        flex.loc[fidx, "MatchRef"] = ",".join(grp["row_id"].astype(str).tolist())
        flex.loc[fidx, "BancoFecha"] = fecha_b
        flex.loc[fidx, "BancoConcepto"] = "Regularización QR período anterior (acumulado)"
        flex.loc[fidx, "BancoComprobante"] = "AGREGADO_QR"
        flex.loc[fidx, "FechaAcreditacionUsada"] = flex.loc[fidx, "FechaFlexxus"]
        flex.loc[fidx, "Diagnostico"] = f"OK - PAV QR agregado contra {len(grp)} CR DEBIN SPOT"

        bank.loc[grp.index, "Matched"] = True
        bank.loc[grp.index, "MatchStage"] = "CORRIENTE_AGREGADO_QR"
        bank.loc[grp.index, "MatchRef"] = flex.loc[fidx, "row_id"]
        bank.loc[grp.index, "FlexxusNumero"] = flex.loc[fidx, "Numero"]
        bank.loc[grp.index, "Diagnostico"] = f"Incluido en PAV QR agregado {flex.loc[fidx, 'Numero']}"

        rows.append({
            "flex_index": fidx,
            "Fecha Flexxus": flex.loc[fidx, "FechaFlexxus"],
            "Número Flexxus": flex.loc[fidx, "Numero"],
            "Categoría": "QR",
            "Concepto Flexxus": flex.loc[fidx, "Movimiento"],
            "Monto Flexxus": monto,
            "Cantidad líneas banco": len(grp),
            "Total banco": total,
            "Diferencia": round(monto - total, 2),
            "Diagnóstico": "OK agregado PAV/QR",
        })
    out = pd.DataFrame(rows)
    if not out.empty and "flex_index" in out.columns:
        out = out.drop(columns=["flex_index"])
    return flex, bank, out



def match_pedidosya_current(flex: pd.DataFrame, bank: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """Match especial Pedidos Ya contra transferencia entrante bancaria.

    Pedidos Ya no es QR ni tarjeta, pero puede acreditarse como transferencia
    entrante con diferencia menor por redondeo. V4.4 permite hasta $1,00 y
    deja diagnóstico explícito sin cupón QR ni número de liquidación.
    """
    rows = []
    flex_idx = flex[
        (~flex["Matched"]) & (~flex["ConsumedPrev"]) &
        (flex["Tipo"] == "PAV") & (flex["EsPedidosYa"])
    ].index.tolist()
    for fidx in flex_idx:
        monto = float(flex.loc[fidx, "MontoFlexxus"])
        fecha_f = flex.loc[fidx, "FechaFlexxus_dt"]
        candidates = bank[
            (~bank["Matched"]) & (~bank["ConsumedPrev"]) &
            (bank["EsIngreso"]) & (bank["Categoria"] == "TRANSFERENCIA_ENTRANTE") &
            (bank["ImporteAbs"].apply(lambda x: abs(float(x) - monto) <= 1.00))
        ].copy()
        if candidates.empty:
            continue
        candidates["date_diff"] = (candidates["Fecha_dt"] - fecha_f).abs()
        bidx = candidates.sort_values(["date_diff", "row_id"]).index[0]
        diff = round(float(bank.loc[bidx, "ImporteAbs"]) - monto, 2)
        flex.loc[fidx, "Matched"] = True
        flex.loc[fidx, "MatchStage"] = "CORRIENTE_PEDIDOSYA"
        flex.loc[fidx, "MatchRef"] = bank.loc[bidx, "row_id"]
        flex.loc[fidx, "BancoFecha"] = bank.loc[bidx, "Fecha"]
        flex.loc[fidx, "BancoConcepto"] = bank.loc[bidx, "Concepto"]
        flex.loc[fidx, "BancoComprobante"] = bank.loc[bidx, "Comprobante"]
        flex.loc[fidx, "FechaAcreditacionUsada"] = bank.loc[bidx, "Fecha"] if bank.loc[bidx, "Fecha_dt"] >= fecha_f else flex.loc[fidx, "FechaFlexxus"]
        flex.loc[fidx, "Diagnostico"] = f"OK - Pedidos Ya contra transferencia entrante; diferencia redondeo {diff:.2f}. No QR/no tarjeta"
        bank.loc[bidx, "Matched"] = True
        bank.loc[bidx, "MatchStage"] = "CORRIENTE_PEDIDOSYA"
        bank.loc[bidx, "MatchRef"] = flex.loc[fidx, "row_id"]
        bank.loc[bidx, "FlexxusNumero"] = flex.loc[fidx, "Numero"]
        bank.loc[bidx, "Diagnostico"] = f"Incluido en PAV Pedidos Ya {flex.loc[fidx, 'Numero']}; diferencia {diff:.2f}"
        rows.append({
            "Fecha Flexxus": flex.loc[fidx, "FechaFlexxus"],
            "Número Flexxus": flex.loc[fidx, "Numero"],
            "Categoría": "PEDIDOS_YA",
            "Concepto Flexxus": flex.loc[fidx, "Movimiento"],
            "Monto Flexxus": monto,
            "Fecha banco": bank.loc[bidx, "Fecha"],
            "Concepto banco": bank.loc[bidx, "Concepto"],
            "Importe banco": float(bank.loc[bidx, "ImporteAbs"]),
            "Diferencia": diff,
            "Diagnóstico": "OK Pedidos Ya / transferencia entrante",
        })
    return flex, bank, pd.DataFrame(rows)

def match_current(flex: pd.DataFrame, bank: pd.DataFrame, qr: pd.DataFrame, trx: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """Matchea movimientos corrientes Flexxus vs banco, sin tocar regularizaciones anteriores."""
    # V4.8: si el archivo actual contiene movimientos Flexxus con fecha anterior
    # al primer movimiento bancario del extracto actual, se consideran filas repetidas
    # del estado anterior de Flexxus. No deben volver a entrar como C1/C2 corriente.
    try:
        bank_min = bank["Fecha_dt"].dropna().min()
        if pd.notna(bank_min) and bool(flex.attrs.get("has_previous_conciliation", False)):
            old_flex_idx = flex[
                (~flex["Matched"]) & (~flex["ConsumedPrev"]) &
                (flex["FechaFlexxus_dt"].notna()) &
                (flex["FechaFlexxus_dt"] < bank_min)
            ].index
            for fidx in old_flex_idx:
                flex.loc[fidx, "ConsumedPrev"] = True
                flex.loc[fidx, "MatchStage"] = "FLEXXUS_ANTERIOR_REPETIDO"
                flex.loc[fidx, "Diagnostico"] = "Fila de Flexxus anterior repetida en reproceso; no entra como C1/C2 corriente"
    except Exception:
        pass
    flex, bank, mbext_ag = match_mbext_current_aggregated(flex, bank)
    flex.attrs["mbext_agregado"] = mbext_ag
    flex, bank, pedidosya_ag = match_pedidosya_current(flex, bank)
    flex.attrs["pedidosya_agregado"] = pedidosya_ag
    flex.attrs["pav_qr_agregado"] = pd.DataFrame()
    # Lookups auxiliares
    qr_lookup = {}
    if not qr.empty and "NetoQR" in qr.columns:
        for _, r in qr.iterrows():
            qr_lookup.setdefault(round(float(r.get("NetoQR", 0)), 2), []).append(r)
    trx_lookup = {}
    if not trx.empty and "MontoNeto" in trx.columns:
        for _, r in trx.iterrows():
            trx_lookup.setdefault(round(float(r.get("MontoNeto", 0)), 2), []).append(r)

    def do_match(tipo: str, bank_side: str):
        flex_idx = flex[(~flex["Matched"]) & (~flex["ConsumedPrev"]) & (flex["Tipo"] == tipo)].index.tolist()
        for idx in flex_idx:
            monto = float(flex.loc[idx, "MontoFlexxus"])
            fecha_f = flex.loc[idx, "FechaFlexxus_dt"]
            if bank_side == "INGRESO":
                candidates = bank[(~bank["Matched"]) & (~bank["ConsumedPrev"]) & bank["EsIngreso"]].copy()
            else:
                candidates = bank[(~bank["Matched"]) & (~bank["ConsumedPrev"]) & bank["EsEgreso"]].copy()
            candidates = candidates[candidates["ImporteAbs"].apply(lambda x: amount_exact(x, monto))]
            if candidates.empty:
                continue
            candidates["date_diff"] = (candidates["Fecha_dt"] - fecha_f).abs()
            best = candidates.sort_values(["date_diff", "row_id"]).iloc[0]
            bidx = best.name

            flex.loc[idx, "Matched"] = True
            flex.loc[idx, "MatchStage"] = "CORRIENTE"
            flex.loc[idx, "MatchRef"] = bank.loc[bidx, "row_id"]
            flex.loc[idx, "BancoFecha"] = bank.loc[bidx, "Fecha"]
            flex.loc[idx, "BancoConcepto"] = bank.loc[bidx, "Concepto"]
            flex.loc[idx, "BancoComprobante"] = bank.loc[bidx, "Comprobante"]
            flex.loc[idx, "Diagnostico"] = "OK corriente - match exacto por importe"
            # Regla fecha acreditación
            if bank.loc[bidx, "Fecha_dt"] < fecha_f:
                flex.loc[idx, "FechaAcreditacionUsada"] = flex.loc[idx, "FechaFlexxus"]
            else:
                flex.loc[idx, "FechaAcreditacionUsada"] = bank.loc[bidx, "Fecha"]

            bank.loc[bidx, "Matched"] = True
            bank.loc[bidx, "MatchStage"] = "CORRIENTE"
            bank.loc[bidx, "MatchRef"] = flex.loc[idx, "row_id"]
            bank.loc[bidx, "FlexxusNumero"] = flex.loc[idx, "Numero"]
            bank.loc[bidx, "Diagnostico"] = "OK corriente - match exacto por importe"

    do_match("PAV", "INGRESO")
    flex, bank, pav_qr_ag = match_pav_qr_current_aggregated(flex, bank)
    flex.attrs["pav_qr_agregado"] = pav_qr_ag
    do_match("MB-ENT-EX", "EGRESO")
    do_match("MB-EXT", "EGRESO")

    return flex, bank

# =============================================================================
# RESULTADOS
# =============================================================================

def get_saldo_flexxus(flex: pd.DataFrame) -> float:
    # Mantiene criterio del sistema anterior: saldo de la serie PAV si existe; si no, último saldo general.
    pav = flex[(flex["Tipo"] == "PAV") & (flex["SaldoFlexxus"].abs() > 0)]
    if not pav.empty:
        return float(pav.sort_values("FechaFlexxus_dt").iloc[-1]["SaldoFlexxus"])
    nonzero = flex[flex["SaldoFlexxus"].abs() > 0]
    if not nonzero.empty:
        return float(nonzero.sort_values("FechaFlexxus_dt").iloc[-1]["SaldoFlexxus"])
    return 0.0


def compute_results(flex: pd.DataFrame, bank: pd.DataFrame, prev_status: pd.DataFrame, regs: pd.DataFrame, mode: str, prev_summary: Optional[Dict[str, float]] = None) -> Dict:
    saldo_f = get_saldo_flexxus(flex)
    saldo_b = get_saldo_banco_final(bank)
    continuity = build_continuity_control(prev_summary or {}, bank)

    # V5.5: ningún movimiento "pendiente anterior repetido" puede quedar en zona muerta.
    # Si vuelve a aparecer en el extracto/archivo actual y sigue abierto, debe figurar en su bloque natural.
    repeat_b = bank[(bank["ConsumedPrev"]) & (bank["MatchStage"].astype(str).eq("PENDIENTE_ANTERIOR_REPETIDO"))].copy()
    base_f = flex[(~flex["Matched"]) & (~flex["ConsumedPrev"])].copy()
    base_b = bank[(~bank["Matched"]) & (~bank["ConsumedPrev"])].copy()
    current_unmatched_f = base_f
    if not repeat_b.empty:
        base_b.attrs = {}
        repeat_b.attrs = {}
        current_unmatched_b = pd.concat([base_b, repeat_b], ignore_index=False)
    else:
        current_unmatched_b = base_b

    if prev_status is None or prev_status.empty:
        prev_open = pd.DataFrame()
    else:
        prev_open = prev_status[~prev_status.get("Regularizado", False)].copy()

    next_pendings = []
    if not prev_open.empty:
        for _, p in prev_open.iterrows():
            d = p.to_dict()
            d["Estado"] = d.get("Estado", "ABIERTO_ANTERIOR") or "ABIERTO_ANTERIOR"
            d["Fuente"] = d.get("Fuente", "") or "Pendiente anterior aún abierto"
            next_pendings.append(d)

    for _, f in current_unmatched_f.iterrows():
        sign = -1 if f["Tipo"] == "PAV" else 1
        next_pendings.append({
            "pending_id": f"P-{uuid.uuid4().hex[:10]}",
            "FechaOrigen": f["FechaFlexxus"],
            "Origen": "FLEXXUS_NO_BANCO",
            "TipoPendiente": "FLEXXUS_INGRESO_NO_BANCO" if f["Tipo"] == "PAV" else "FLEXXUS_EGRESO_NO_BANCO",
            "TipoMovimiento": f["Tipo"],
            "Numero": f["Numero"],
            "Concepto": f["Movimiento"],
            "Categoria": "PAV" if f["Tipo"] == "PAV" else f["Tipo"],
            "Monto": float(f["MontoFlexxus"]),
            "SignoCalculo": sign,
            "Estado": "ABIERTO_NUEVO",
            "Fuente": "Corriente actual",
        })
    for _, b in current_unmatched_b.iterrows():
        sign = 1 if b["EsIngreso"] else -1
        next_pendings.append({
            "pending_id": f"P-{uuid.uuid4().hex[:10]}",
            "FechaOrigen": b["Fecha"],
            "Origen": "BANCO_NO_FLEXXUS",
            "TipoPendiente": "BANCO_INGRESO_NO_FLEXXUS" if b["EsIngreso"] else "BANCO_EGRESO_NO_FLEXXUS",
            "TipoMovimiento": "PAV" if b["EsIngreso"] else ("MB-EXT" if b["Categoria"] in IMP_CATS else "MB-ENT-EX"),
            "Numero": b["Comprobante"],
            "Concepto": b["Concepto"],
            "Categoria": b["Categoria"],
            "Monto": float(b["ImporteAbs"]),
            "SignoCalculo": sign,
            "Estado": "ABIERTO_NUEVO",
            "Fuente": "Corriente actual",
        })

    pend_next = assign_stable_pending_ids(pd.DataFrame(next_pendings))
    if pend_next.empty:
        C1 = C2 = C3 = C4 = 0.0
    else:
        pf = pend_next[pend_next["Origen"].astype(str).eq("FLEXXUS_NO_BANCO")].copy()
        pbi = pend_next[(pend_next["Origen"].astype(str).eq("BANCO_NO_FLEXXUS")) & (pend_next["SignoCalculo"].astype(float) > 0)].copy()
        pbe = pend_next[(pend_next["Origen"].astype(str).eq("BANCO_NO_FLEXXUS")) & (pend_next["SignoCalculo"].astype(float) < 0)].copy()
        C1 = float(pf[pf["TipoMovimiento"].astype(str).eq("PAV")]["Monto"].sum()) if not pf.empty else 0.0
        C2 = float(pf[~pf["TipoMovimiento"].astype(str).eq("PAV")]["Monto"].sum()) if not pf.empty else 0.0
        C3 = float(pbi["Monto"].sum()) if not pbi.empty else 0.0
        C4 = float(pbe["Monto"].sum()) if not pbe.empty else 0.0

    calc = saldo_f - C1 + C2 + C3 - C4
    diff = round(saldo_b - calc, 2)

    if abs(diff) > 0.005 and abs(diff) <= 1.0:
        fecha_aj = fmt_date(bank["Fecha_dt"].max()) if not bank.empty and "Fecha_dt" in bank.columns else ""
        ajuste_row = {
            "row_id": f"AJ-{uuid.uuid4().hex[:10]}",
            "SourceOrder": 999999,
            "Fecha": fecha_aj,
            "Fecha_dt": safe_date(fecha_aj),
            "Comprobante": "CTRL-CONTINUIDAD",
            "Concepto": "Diferencia técnica menor de continuidad/extracto bancario",
            "ConceptoNorm": "DIFERENCIA TECNICA MENOR CONTINUIDAD EXTRACTO BANCARIO",
            "ImporteNum": abs(diff) if diff > 0 else -abs(diff),
            "ImporteAbs": abs(diff),
            "SaldoNum": saldo_b,
            "EsIngreso": diff > 0,
            "EsEgreso": diff < 0,
            "Categoria": "AJUSTE_MENOR_REDONDEO",
            "Matched": False,
            "ConsumedPrev": False,
            "MatchStage": "CONTROL_CONTINUIDAD_MENOR",
            "MatchRef": "",
            "FlexxusNumero": "",
            "Diagnostico": "Control visible: diferencia menor <= $1 entre fórmula y extracto; no es ajuste oculto",
        }
        current_unmatched_b = pd.concat([current_unmatched_b, pd.DataFrame([ajuste_row])], ignore_index=True)
        next_pendings.append({
            "pending_id": stable_pending_id("BANCO_NO_FLEXXUS", fecha_aj, "CTRL-CONTINUIDAD", "CTRL-CONTINUIDAD", "Diferencia técnica menor de continuidad/extracto bancario", "AJUSTE_MENOR_REDONDEO", abs(diff)),
            "FechaOrigen": fecha_aj,
            "Origen": "BANCO_NO_FLEXXUS",
            "TipoPendiente": "BANCO_INGRESO_NO_FLEXXUS" if diff > 0 else "BANCO_EGRESO_NO_FLEXXUS",
            "TipoMovimiento": "PAV" if diff > 0 else "MB-EXT",
            "Numero": "CTRL-CONTINUIDAD",
            "Concepto": "Diferencia técnica menor de continuidad/extracto bancario",
            "Categoria": "AJUSTE_MENOR_REDONDEO",
            "Monto": abs(float(diff)),
            "SignoCalculo": 1 if diff > 0 else -1,
            "Estado": "CONTROL_CONTINUIDAD_MENOR",
            "Fuente": "Control de cobertura",
        })
        pend_next = assign_stable_pending_ids(pd.DataFrame(next_pendings))
        if diff > 0:
            C3 += abs(diff)
        else:
            C4 += abs(diff)
        calc = saldo_f - C1 + C2 + C3 - C4
        diff = round(saldo_b - calc, 2)

    return {
        "mode": mode,
        "saldo_flexxus": float(saldo_f),
        "saldo_banco": float(saldo_b),
        "C1": float(C1),
        "C2": float(C2),
        "C3": float(C3),
        "C4": float(C4),
        "prev_open_effect": 0.0,
        "calc_final": float(calc),
        "diferencia": float(diff),
        "current_unmatched_f": current_unmatched_f.copy(),
        "current_unmatched_b": current_unmatched_b.copy(),
        "prev_open": prev_open.copy() if not prev_open.empty else pd.DataFrame(),
        "regularizaciones": regs.copy() if regs is not None else pd.DataFrame(),
        "mbext_agregado": flex.attrs.get("mbext_agregado", pd.DataFrame()),
        "pav_qr_agregado": flex.attrs.get("pav_qr_agregado", pd.DataFrame()),
        "continuity": continuity,
        "pendientes_proxima": pend_next,
        "matched_flex": flex[flex["Matched"]].copy(),
        "matched_bank": bank[bank["Matched"]].copy(),
        "consumed_prev_flex": flex[flex["ConsumedPrev"]].copy(),
        "consumed_prev_bank": bank[bank["ConsumedPrev"]].copy(),
    }



def split_pendientes_para_ui(pend: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """Separa la tabla interna de trazabilidad en tres grupos visibles.

    - Pendientes reales: movimientos abiertos que siguen vivos para próxima conciliación.
    - Procesados históricos: filas usadas solo para trazabilidad/reproceso, no son pendientes.
    - Controles: ajustes técnicos menores o controles de continuidad visibles.
    """
    if pend is None or pend.empty:
        empty = pd.DataFrame()
        return empty, empty, empty

    df = pend.copy()
    for col in ["Origen", "Estado", "TipoPendiente", "Categoria"]:
        if col not in df.columns:
            df[col] = ""
    if "SignoCalculo" not in df.columns:
        df["SignoCalculo"] = 0

    origen = df["Origen"].astype(str)
    estado = df["Estado"].astype(str)
    tipo = df["TipoPendiente"].astype(str)
    categoria = df["Categoria"].astype(str)
    signo = pd.to_numeric(df["SignoCalculo"], errors="coerce").fillna(0.0)

    is_procesado = (
        origen.str.contains("PROCESADO", case=False, na=False)
        | estado.str.contains("PROCESADO", case=False, na=False)
        | tipo.str.contains("PROCESADO", case=False, na=False)
    )
    is_control = (
        estado.str.contains("CONTROL", case=False, na=False)
        | categoria.eq("AJUSTE_MENOR_REDONDEO")
        | df.get("Numero", pd.Series([""] * len(df), index=df.index)).astype(str).str.contains("CTRL", case=False, na=False)
    )
    is_natural_open = origen.isin(["FLEXXUS_NO_BANCO", "BANCO_NO_FLEXXUS"]) & (signo != 0)

    pendientes_reales = df[is_natural_open & ~is_procesado & ~is_control].copy()
    procesados = df[is_procesado].copy()
    controles = df[is_control & ~is_procesado].copy()
    return pendientes_reales, procesados, controles

# =============================================================================
# EXCEL OUTPUT
# =============================================================================

def write_df(ws, start_row: int, start_col: int, df: pd.DataFrame, money_cols: Optional[List[str]] = None, date_cols: Optional[List[str]] = None):
    money_cols = money_cols or []
    date_cols = date_cols or []
    hdr_fill = PatternFill("solid", fgColor="1F4E78")
    hdr_font = Font(bold=True, color="FFFFFF")
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for j, col in enumerate(df.columns, start_col):
        c = ws.cell(start_row, j, col)
        c.fill = hdr_fill
        c.font = hdr_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border
    for i, (_, row) in enumerate(df.iterrows(), start_row + 1):
        for j, col in enumerate(df.columns, start_col):
            val = row[col]
            if isinstance(val, (np.integer, np.floating)):
                val = float(val)
            c = ws.cell(i, j, val)
            c.border = border
            c.alignment = Alignment(vertical="top", wrap_text=True)
            if col in money_cols:
                c.number_format = '#,##0.00'
            if col in date_cols:
                c.number_format = "dd/mm/yyyy"
    for col_idx in range(start_col, start_col + len(df.columns)):
        letter = get_column_letter(col_idx)
        max_len = 10
        for cell in ws[letter]:
            max_len = max(max_len, len(str(cell.value or "")))
        ws.column_dimensions[letter].width = min(max_len + 2, 42)
    ws.freeze_panes = ws.cell(start_row + 1, start_col).coordinate
    try:
        ws.auto_filter.ref = ws.dimensions
    except Exception:
        pass



def build_excel_report(flex: pd.DataFrame, bank: pd.DataFrame, qr: pd.DataFrame, trx: pd.DataFrame, res: Dict) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Conciliacion semanal"

    hdr_fill = PatternFill("solid", fgColor="4472C4")
    hdr_f = Font(bold=True, size=10, name="Calibri", color="FFFFFF")
    norm = Font(size=10, name="Calibri")
    bold_f = Font(bold=True, size=10, name="Calibri")
    title_f = Font(bold=True, size=14, name="Calibri")
    sub_f = Font(bold=True, size=12, name="Calibri")
    grn_fill = PatternFill("solid", fgColor="E2EFDA")
    red_fill = PatternFill("solid", fgColor="FCE4EC")
    money_fmt = "#,##0.00"
    bdr = Border(
        left=Side("thin", color="D9E2F3"),
        right=Side("thin", color="D9E2F3"),
        top=Side("thin", color="D9E2F3"),
        bottom=Side("thin", color="D9E2F3"),
    )

    def hw(sheet, row, headers):
        for c, h in enumerate(headers, 1):
            x = sheet.cell(row, c, h)
            x.font = hdr_f
            x.fill = hdr_fill
            x.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            x.border = bdr

    def dw(sheet, row, vals, mc=None):
        mc = mc or []
        for c, v in enumerate(vals, 1):
            x = sheet.cell(row, c, v)
            x.font = norm
            x.border = bdr
            x.alignment = Alignment(wrap_text=True, vertical="top")
            if c in mc:
                x.number_format = money_fmt

    def aw(sheet, mx=45):
        for col in sheet.columns:
            cl = get_column_letter(col[0].column)
            ml = max((len(str(c.value or "")) for c in col), default=8)
            sheet.column_dimensions[cl].width = min(max(ml + 3, 10), mx)

    def frz(sheet):
        try:
            sheet.auto_filter.ref = sheet.dimensions
            sheet.freeze_panes = "A2"
        except Exception:
            pass

    def safe_min_date(df, col):
        if df is None or df.empty or col not in df.columns:
            return ""
        v = df[col].dropna()
        if v.empty:
            return ""
        return fmt_date(v.min())

    def safe_max_date(df, col):
        if df is None or df.empty or col not in df.columns:
            return ""
        v = df[col].dropna()
        if v.empty:
            return ""
        return fmt_date(v.max())

    period_min = safe_min_date(bank, "Fecha_dt")
    period_max = safe_max_date(bank, "Fecha_dt")
    titulo_periodo = f"Conciliación al {period_max} (período {period_min} al {period_max}; banco con movimientos hasta {period_max})"

    pend = res.get("pendientes_proxima", pd.DataFrame()).copy()
    if pend is None:
        pend = pd.DataFrame()

    def pend_flex():
        if pend.empty:
            return pd.DataFrame()
        return pend[pend["Origen"].astype(str).eq("FLEXXUS_NO_BANCO")].copy()

    def pend_banco_ing():
        if pend.empty:
            return pd.DataFrame()
        return pend[(pend["Origen"].astype(str).eq("BANCO_NO_FLEXXUS")) & (pend["SignoCalculo"].astype(float) > 0)].copy()

    def pend_banco_egr():
        if pend.empty:
            return pd.DataFrame()
        return pend[(pend["Origen"].astype(str).eq("BANCO_NO_FLEXXUS")) & (pend["SignoCalculo"].astype(float) < 0)].copy()

    pf = pend_flex()
    pbi = pend_banco_ing()
    pbe = pend_banco_egr()

    def ar_money(v):
        try:
            txt = f"{float(v):,.2f}"
            return txt.replace(",", "X").replace(".", ",").replace("X", ".")
        except Exception:
            return str(v)

    def canonical_flexxus_movement(cat, concepto=""):
        c = norm_txt(str(concepto))
        if cat == "RETENCION_IIBB":
            return "INGRESOS BRUTOS MENDOZA"
        if cat == "IMPUESTO_LEY25413_DEB":
            return "GRAVAMEN LEY 25.413 S/DEB"
        if cat == "IMPUESTO_LEY25413_CRED":
            return "GRAVAMEN LEY 25.413 S/CRED"
        if cat == "IVA":
            return "I.V.A. BASE"
        if cat == "GASTO_BANCARIO":
            return "COM TRANSFE ELECTRONICA"
        if cat == "DEBITO_LIQ_TARJETA":
            return "DEB LIQ MASTERCARD 24H"
        if cat == "DEBITO_PAGO_DIRECTO":
            return "DEBITO PAGO DIRECTO"
        return str(concepto).strip() or CAT_LABELS.get(cat, cat)

    def flexxus_copy_instruction(cat, monto, fecha="", concepto=""):
        tipo = "MB-EXT" if cat in IMP_CATS else "MB-ENT-EX"
        mov = canonical_flexxus_movement(cat, concepto)
        fecha_txt = f" | Fecha: {fecha}" if fecha else ""
        return f"Copiar en Flexxus: Tipo={tipo} | Movimiento={mov} | Monto={ar_money(monto)}{fecha_txt}"

    def enrich_flexxus_pending_row(p):
        monto = float(p.get("Monto", 0.0) or 0.0)
        concepto = str(p.get("Concepto", ""))
        tipo = str(p.get("TipoMovimiento", ""))
        c = norm_txt(concepto)
        if "PEDIDOS YA" in c:
            return "", "", "", "", ""

        best_qr = None
        best_qr_diff = float("inf")
        if qr is not None and not qr.empty and "NetoQR" in qr.columns:
            for _, q in qr.iterrows():
                try:
                    neto = float(q.get("NetoQR", 0.0) or 0.0)
                except Exception:
                    continue
                diff = abs(neto - monto)
                if diff < best_qr_diff and diff <= max(1.0, monto * 0.02):
                    best_qr = q
                    best_qr_diff = diff

        best_trx = None
        best_trx_diff = float("inf")
        if trx is not None and not trx.empty and "MontoNeto" in trx.columns:
            for _, t in trx.iterrows():
                try:
                    neto = float(t.get("MontoNeto", 0.0) or 0.0)
                except Exception:
                    continue
                diff = abs(neto - monto)
                if diff < best_trx_diff and diff <= max(1.0, monto * 0.02):
                    best_trx = t
                    best_trx_diff = diff

        prefer_trx = ("TARJETA" in c or "CREDITO" in c or "CRÉDITO" in c or tipo == "PAV") and best_trx is not None
        if prefer_trx or (best_trx is not None and (best_qr is None or best_trx_diff <= best_qr_diff)):
            comercio = str(best_trx.get("COMERCIO", ""))
            local = LOCAL_MAP.get(comercio, str(best_trx.get("Local", comercio)))
            liq = str(best_trx.get("NUMERO LIQUIDACION", best_trx.get("Número Liquidación", "")))
            comp = round(float(best_trx.get("MontoNeto", 0.0) or 0.0), 2)
            return local, "", liq, comp, round(monto - comp, 2)

        if best_qr is not None:
            cod = str(best_qr.get("CodComercio", best_qr.get("Cód. comercio", "")))
            local = LOCAL_MAP.get(cod, cod)
            cupon = str(best_qr.get("Cupon", best_qr.get("Ticket", best_qr.get("Id QR", ""))))
            comp = round(float(best_qr.get("NetoQR", 0.0) or 0.0), 2)
            return local, cupon, "", comp, round(monto - comp, 2)

        if "TARJETA" in c or "CREDITO" in c or "CRÉDITO" in c:
            return "—", "", "No encontrado en TRX", "—", ""
        return "—", "No encontrado en QR ni TRX", "", "—", ""

    def enrich_bank_income_pending_row(p):
        """Identifica local/cupón/liquidación para ingresos de Banco no Flexxus.

        V5.5:
        1) Si el comprobante bancario es un código de comercio Merchant, asigna local directo.
        2) Si es liquidación de tarjeta, busca match exacto en TRX por importe y comercio.
        3) Si es QR, busca match exacto en QR PCT por neto e informa local/cupón.
        4) Si no hay match fino, nunca deja local vacío cuando el banco trae comercio.
        """
        monto = float(p.get("Monto", 0.0) or 0.0)
        cat = str(p.get("Categoria", ""))
        concepto = str(p.get("Concepto", ""))
        comprobante = str(p.get("Numero", p.get("Comprobante banco", ""))).strip()
        cat_norm = norm_txt(cat + " " + concepto)
        comp_digits = re.sub(r"[^0-9]", "", comprobante)
        local_from_bank = LOCAL_MAP.get(comp_digits, "")

        is_card = (cat == "LIQUIDACION_TARJETA" or "LIQ" in cat_norm or "TARJETA" in cat_norm or "AMEX" in cat_norm or "MASTER" in cat_norm)
        is_qr = (cat == "QR" or "DEBIN" in cat_norm or "QR" in cat_norm)

        best_trx = None
        best_trx_diff = float("inf")
        if is_card and trx is not None and not trx.empty and "MontoNeto" in trx.columns:
            trx_candidates = trx.copy()
            if comp_digits and "COMERCIO" in trx_candidates.columns:
                by_comercio = trx_candidates[trx_candidates["COMERCIO"].astype(str).str.replace(r"\D", "", regex=True).eq(comp_digits)]
                if not by_comercio.empty:
                    trx_candidates = by_comercio
            for _, t in trx_candidates.iterrows():
                try:
                    neto = float(t.get("MontoNeto", t.get("IMPORTE NETO", 0.0)) or 0.0)
                except Exception:
                    continue
                diff = abs(neto - monto)
                if diff < best_trx_diff and diff <= max(0.05, monto * 0.00001):
                    best_trx = t
                    best_trx_diff = diff

        if best_trx is not None:
            comercio = re.sub(r"[^0-9]", "", str(best_trx.get("COMERCIO", "")))
            local = LOCAL_MAP.get(comercio, local_from_bank or str(best_trx.get("Local", comercio)) or "No identificado")
            liq = str(best_trx.get("NUMERO LIQUIDACION", best_trx.get("Número Liquidación", "")))
            comp = round(float(best_trx.get("MontoNeto", best_trx.get("IMPORTE NETO", 0.0)) or 0.0), 2)
            return local, "", liq, comp, round(monto - comp, 2), "TRX Merchant", "Identificado por TRX / liquidación tarjeta"

        best_qr = None
        best_qr_diff = float("inf")
        if is_qr and qr is not None and not qr.empty and "NetoQR" in qr.columns:
            for _, q in qr.iterrows():
                try:
                    neto = float(q.get("NetoQR", 0.0) or 0.0)
                except Exception:
                    continue
                diff = abs(neto - monto)
                if diff < best_qr_diff and diff <= max(0.05, monto * 0.00001):
                    best_qr = q
                    best_qr_diff = diff

        if best_qr is not None:
            cod = re.sub(r"[^0-9]", "", str(best_qr.get("CodComercio", best_qr.get("Cód. comercio", ""))))
            local = LOCAL_MAP.get(cod, local_from_bank or cod or "No identificado")
            cupon = str(best_qr.get("Cupon", best_qr.get("Ticket", best_qr.get("Id QR", ""))))
            comp = round(float(best_qr.get("NetoQR", 0.0) or 0.0), 2)
            return local, cupon, "", comp, round(monto - comp, 2), "QR PCT", "Identificado por Transacciones QR"

        if local_from_bank and is_card:
            return local_from_bank, "", "No encontrado en TRX", "—", "", "Banco/comprobante", "Local identificado por comprobante bancario; falta número de liquidación TRX"
        if local_from_bank:
            return local_from_bank, "", "", "—", "", "Banco/comprobante", "Local identificado por comprobante bancario"

        if cat == "TRANSFERENCIA_ENTRANTE":
            return "No identificado", "", "", "—", "", "Banco", "Transferencia entrante sin local detectado"
        if is_card:
            return "No identificado", "", "No encontrado en TRX", "—", "", "Banco", "Liquidación tarjeta sin match TRX/local"
        if is_qr:
            return "No identificado", "No encontrado en QR", "", "—", "", "Banco", "QR bancario sin match QR PCT/local"
        return "No identificado", "", "", "—", "", "Banco", "Origen/local no identificado"

    def add_local_delay_rows(base_rows, local, bloque, categoria, monto, estado, origen, accion):
        loc = str(local or "No identificado").strip() or "No identificado"
        base_rows.append({
            "Local": loc,
            "Bloque": bloque,
            "Categoría": CAT_LABELS.get(categoria, categoria),
            "Cantidad": 1,
            "Importe": float(monto or 0.0),
            "Estado trazabilidad": estado,
            "Origen detectado": origen,
            "Acción sugerida": accion,
        })

    # HOJA 1: Conciliacion semanal
    ws.cell(1, 1, "DANCONA ALIMENTOS - CONCILIACIÓN BANCARIA SEMANAL").font = title_f
    ws.merge_cells("A1:C1")
    ws.cell(2, 1, "Banco Nación Argentina - Mendoza").font = norm
    ws.cell(3, 1, titulo_periodo).font = norm
    ws.merge_cells("A3:C3")

    ws.cell(5, 1, "Concepto").font = bold_f
    ws.cell(5, 2, "Importe").font = bold_f

    resumen_rows = [
        (6, f"Saldo S/Extracto Bancario al {period_max}", float(res["saldo_banco"])),
        (7, f"Saldo S/FLEXXUS al {period_max}", float(res["saldo_flexxus"])),
        (8, "DIFERENCIA INICIAL (extracto banco vs libro Flexxus)", "=B6-B7"),
        (10, "Menos: INGRESOS registrados en FLEXXUS que NO están en el Banco", "=SUMIF('Flexxus no Banco'!D:D,\"PAV\",'Flexxus no Banco'!G:G)"),
        (11, "Más: EGRESOS registrados en FLEXXUS que NO están en el Banco", "=SUM('Flexxus no Banco'!G:G)-B10"),
        (12, "Más: INGRESOS en el Banco pero NO registrados en FLEXXUS", "=SUM('Banco ingresos no Flexxus'!H:H)"),
        (13, "Menos: EGRESOS en el Banco pero NO registrados en FLEXXUS", "=SUM('Banco egresos no Flexxus'!I:I)"),
        (14, "Pendientes anteriores: incluidos dentro de los bloques anteriores (NO impacta como residual)", "=0"),
        (16, f"SALDO REAL S/FLEXXUS SEGÚN ARCHIVO ACTUAL al {period_max}", "=B7"),
        (17, "SALDO FLEXXUS AJUSTADO PARA CONCILIACIÓN BANCARIA (= Flex - C1 + C2)", "=B7-B10+B11"),
        (18, "TOTAL MOVIMIENTOS CONCILIADOS / .XLS A IMPORTAR", "='Carga Flexxus'!C5"),
        (19, "SALDO BANCO CALCULADO (= Flex - C1 + C2 + C3 - C4)", "=B7-B10+B11+B12-B13"),
        (20, f"SALDO SEGÚN EXTRACTO BANCO al {period_max}", "=B6"),
        (21, "DIFERENCIA FINAL (debe ser 0)", "=B20-B19"),
        (22, "SALDO REAL EN FLEXXUS LUEGO DE IMPORTAR / POST-IMPORT SI APLICA", "=B7"),
        (23, "ACLARACIÓN: B17 es saldo ajustado conciliatorio, no saldo final real", "=B17"),
    ]
    for row, label, val in resumen_rows:
        ws.cell(row, 1, label).font = bold_f if row >= 16 or row in [21] else norm
        c = ws.cell(row, 2, val)
        c.number_format = money_fmt
        if row == 21:
            c.fill = grn_fill if abs(float(res.get("diferencia", 0))) < 1.0 else red_fill
            c.font = Font(bold=True, size=11, name="Calibri", color="375623" if abs(float(res.get("diferencia", 0))) < 1.0 else "C00000")

    ws.cell(25, 1, "Controles de consistencia interna").font = sub_f
    hw(ws, 26, ["Control", "Fórmula / Resultado esperado", "Estado"])
    controles = [
        ("B10 = suma hoja Flexxus no Banco PAV", "=B10-SUMIF('Flexxus no Banco'!D:D,\"PAV\",'Flexxus no Banco'!G:G)", "Debe dar 0"),
        ("B11 = suma hoja Flexxus no Banco egresos", "=B11-(SUM('Flexxus no Banco'!G:G)-B10)", "Debe dar 0"),
        ("B12 = suma hoja Banco ingresos no Flexxus", "=B12-SUM('Banco ingresos no Flexxus'!H:H)", "Debe dar 0"),
        ("B13 = suma hoja Banco egresos no Flexxus", "=B13-SUM('Banco egresos no Flexxus'!I:I)", "Debe dar 0"),
        ("B14 no es residual de cierre", "=B14", "Debe dar 0"),
        ("B18 = total hoja Carga Flexxus", "=B18-'Carga Flexxus'!C5", "Debe dar 0"),
    ]
    rn = 27
    for ctrl, formula, estado in controles:
        dw(ws, rn, [ctrl, formula, estado], mc=[2])
        rn += 1

    rn += 2
    ws.cell(rn, 1, "Detalle de diferencias por bloque").font = sub_f
    hw(ws, rn + 1, ["Bloque", "Fecha", "Detalle / cuenta sugerida", "Cantidad", "Importe", "Observación"])
    rn += 2

    cur_f = res["pendientes_proxima"]
    pf_all = cur_f[cur_f["Origen"].astype(str).eq("FLEXXUS_NO_BANCO")] if not cur_f.empty else pd.DataFrame()
    pbi_all = cur_f[(cur_f["Origen"].astype(str).eq("BANCO_NO_FLEXXUS")) & (cur_f["SignoCalculo"].astype(float) > 0)] if not cur_f.empty else pd.DataFrame()
    pbe_all = cur_f[(cur_f["Origen"].astype(str).eq("BANCO_NO_FLEXXUS")) & (cur_f["SignoCalculo"].astype(float) < 0)] if not cur_f.empty else pd.DataFrame()

    ws.cell(rn, 1, "Menos: INGRESOS registrados en FLEXXUS que NO están en el Banco").font = bold_f
    ws.cell(rn, 5, "=B10").number_format = money_fmt
    rn += 1
    if not pf_all.empty:
        pav_all = pf_all[pf_all["TipoMovimiento"].astype(str).eq("PAV")]
        if not pav_all.empty:
            for fecha, grp in pav_all.groupby("FechaOrigen"):
                dw(ws, rn, ["Flexxus ingreso no banco", fecha, f"{len(grp)} PAV sin acreditar en banco", len(grp), float(grp["Monto"].sum()), "Detalle en hoja Flexxus no Banco"], mc=[5])
                rn += 1

    ws.cell(rn, 1, "Más: EGRESOS registrados en FLEXXUS que NO están en el Banco").font = bold_f
    ws.cell(rn, 5, "=B11").number_format = money_fmt
    rn += 1
    if not pf_all.empty:
        fegr_all = pf_all[~pf_all["TipoMovimiento"].astype(str).eq("PAV")]
        if not fegr_all.empty:
            for _, f in fegr_all.iterrows():
                dw(ws, rn, ["Flexxus egreso no banco", f.get("FechaOrigen", ""), f"{f.get('TipoMovimiento','')}: {str(f.get('Concepto',''))[:40]}", 1, float(f.get("Monto", 0.0)), f.get("Estado", "")], mc=[5])
                rn += 1

    ws.cell(rn, 1, "Más: INGRESOS en el Banco pero NO registrados en FLEXXUS").font = bold_f
    ws.cell(rn, 5, "=B12").number_format = money_fmt
    rn += 1
    if not pbi_all.empty:
        for cat, grp in pbi_all.groupby("Categoria"):
            label = CAT_LABELS.get(cat, cat)
            dw(ws, rn, ["Banco ingreso no Flexxus", "", label, len(grp), float(grp["Monto"].sum()), "Detalle en hoja Banco ingresos no Flexxus"], mc=[5])
            rn += 1

    ws.cell(rn, 1, "Menos: EGRESOS en el Banco pero NO registrados en FLEXXUS").font = bold_f
    ws.cell(rn, 5, "=B13").number_format = money_fmt
    rn += 1
    if not pbe_all.empty:
        for cat, grp in pbe_all.groupby("Categoria"):
            label = CAT_LABELS.get(cat, cat)
            total_cat = float(grp["Monto"].sum())
            obs = flexxus_copy_instruction(cat, total_cat, concepto=str(grp.iloc[0].get("Concepto", ""))) if cat in IMP_CATS else "Revisar según concepto antes de cargar"
            dw(ws, rn, ["Banco egreso no Flexxus", "", label, len(grp), total_cat, obs], mc=[5])
            rn += 1

    ws.cell(rn, 1, "Pendientes anteriores").font = bold_f
    ws.cell(rn, 5, "=B14").number_format = money_fmt
    dw(ws, rn + 1, ["Info", "", "Los pendientes anteriores abiertos están dentro de Flexxus no Banco, Banco ingresos no Flexxus o Banco egresos no Flexxus. No se usa residual.", "", 0, "No impacta el cierre"], mc=[5])

    ws.column_dimensions["A"].width = 62
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 45
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 45

    # HOJA 2: Resumen banco
    ws_rb = wb.create_sheet("Resumen banco")
    ws_rb.cell(1, 1, "Resumen banco no registrado en Flexxus - apertura para asientos").font = sub_f
    hw(ws_rb, 2, ["Concepto banco", "Cuenta / apertura sugerida", "Tratamiento", "Cantidad", "Importe", "Bloque"])
    rn = 3
    if not pbe.empty:
        for cat, grp in pbe.groupby("Categoria"):
            label = CAT_LABELS.get(cat, cat)
            concepto = str(grp.iloc[0].get("Concepto", cat)).split(" - ")[0]
            total_cat = float(grp["Monto"].sum())
            trat = flexxus_copy_instruction(cat, total_cat, concepto=concepto) if cat in IMP_CATS else "Revisar según concepto"
            dw(ws_rb, rn, [concepto, label, trat, len(grp), total_cat, "Banco egreso no Flexxus"], mc=[5])
            rn += 1
    aw(ws_rb)

    # HOJA 3: Flexxus no Banco
    ws3 = wb.create_sheet("Flexxus no Banco")
    ws3.cell(1, 1, "Detalle de movimientos registrados en Flexxus que no están acreditados en banco").font = sub_f
    hw(ws3, 2, ["ID trazabilidad", "Estado trazabilidad", "Fecha Flexxus", "Tipo", "Número", "Movimiento", "Monto", "Local", "Cupón QR", "Número de Liquidación de Tarjeta", "Monto comparado", "Diferencia"])
    rn = 3
    if not pf.empty:
        for _, p in pf.iterrows():
            local, cupon, liq, monto_comp, dif = enrich_flexxus_pending_row(p)
            dw(ws3, rn, [p.get("pending_id", ""), p.get("Estado", ""), p.get("FechaOrigen", ""), p.get("TipoMovimiento", ""), p.get("Numero", ""), p.get("Concepto", ""), float(p.get("Monto", 0.0)), local, cupon, liq, monto_comp, dif], mc=[7, 11, 12])
            rn += 1
    frz(ws3); aw(ws3)

    # HOJA 4: Banco ingresos no Flexxus
    ws4 = wb.create_sheet("Banco ingresos no Flexxus")
    ws4.cell(1, 1, "Detalle de ingresos del banco no registrados en Flexxus - con diagnóstico de origen").font = sub_f
    hw(ws4, 2, ["ID trazabilidad", "Estado trazabilidad", "Bloque", "Fecha banco", "Comprobante banco", "Concepto banco", "Categoría", "Importe", "Saldo banco", "Origen detectado", "Archivo donde aparece", "Local", "Cupón QR", "Número de Liquidación de Tarjeta", "Monto comparado", "Diferencia", "Diagnóstico", "Acción sugerida"])
    rn = 3
    if not pbi.empty:
        for _, p in pbi.iterrows():
            cat = p.get("Categoria", "")
            cat_label = CAT_LABELS.get(cat, cat)
            local, cupon, liq, monto_comp, dif, archivo_aux, diag_local = enrich_bank_income_pending_row(p)
            archivo = str(p.get("Fuente", ""))
            if archivo_aux and archivo_aux not in archivo:
                archivo = (archivo + " | " + archivo_aux).strip(" |")
            accion = "Cargar/regularizar en Flexxus como PAV o identificar cliente/local"
            if local == "No identificado":
                accion = "URGENTE: identificar local/origen antes de cerrar rutina administrativa"
            dw(ws4, rn, [p.get("pending_id", ""), p.get("Estado", ""), "Banco ingreso no Flexxus", p.get("FechaOrigen", ""), p.get("Numero", ""), p.get("Concepto", ""), cat_label, float(p.get("Monto", 0.0)), "", cat_label, archivo, local, cupon, liq, monto_comp, dif, diag_local, accion], mc=[8, 9, 15, 16])
            rn += 1
    frz(ws4); aw(ws4)

    # HOJA 5: Banco egresos no Flexxus
    ws5 = wb.create_sheet("Banco egresos no Flexxus")
    ws5.cell(1, 1, "Detalle de egresos del banco no registrados en Flexxus").font = sub_f
    hw(ws5, 2, ["ID trazabilidad", "Estado trazabilidad", "Bloque", "Fecha banco", "Comprobante banco", "Concepto banco", "Cuenta / apertura sugerida", "Tratamiento", "Importe", "Saldo banco", "Acción"])
    rn = 3
    if not pbe.empty:
        for _, p in pbe.iterrows():
            cat = p.get("Categoria", "")
            cuenta = CAT_LABELS.get(cat, cat)
            trat = "Retención de Ingresos Brutos" if cat == "RETENCION_IIBB" else ("Anticipo Impuesto a las Ganancias" if "25413" in cat else ("IVA crédito fiscal / revisar" if cat == "IVA" else "Revisar según concepto"))
            accion = flexxus_copy_instruction(cat, float(p.get("Monto", 0.0)), p.get("FechaOrigen", ""), p.get("Concepto", "")) if cat in IMP_CATS else "Registrar asiento / revisar según concepto"
            dw(ws5, rn, [p.get("pending_id", ""), p.get("Estado", ""), "Banco egreso no Flexxus", p.get("FechaOrigen", ""), p.get("Numero", ""), p.get("Concepto", ""), cuenta, trat, float(p.get("Monto", 0.0)), "", accion], mc=[9, 10])
            rn += 1
    frz(ws5); aw(ws5)

    # HOJA 6: Regularizaciones
    ws_reg = wb.create_sheet("Regularizaciones")
    ws_reg.cell(1, 1, "Regularizaciones y ajustes ya incorporados").font = sub_f
    hw(ws_reg, 2, ["ID trazabilidad", "Concepto", "Fecha Flexxus", "Tipo", "Número", "Movimiento", "Cant. banco", "Importe", "Observación"])
    rn = 3
    regs = res.get("regularizaciones", pd.DataFrame())
    if regs is not None and not regs.empty:
        for _, r in regs.iterrows():
            dw(ws_reg, rn, [r.get("ID pendiente", ""), r.get("Origen pendiente", "Regularización período anterior"), r.get("Fecha regularización", ""), r.get("Tipo actual", ""), r.get("Referencia actual", ""), r.get("Concepto actual", ""), "", float(r.get("Monto actual", r.get("Monto pendiente", 0.0))), r.get("Diagnóstico", "")], mc=[8])
            rn += 1
    ag = res.get("mbext_agregado", pd.DataFrame())
    if ag is not None and not ag.empty:
        for _, a in ag.iterrows():
            dw(ws_reg, rn, ["AGREGADO-MBEXT", "MB-EXT agregado contra banco", a.get("Fecha Flexxus", ""), "MB-EXT", a.get("Número Flexxus", ""), a.get("Concepto Flexxus", ""), a.get("Cantidad líneas banco", ""), float(a.get("Monto Flexxus", 0.0)), a.get("Diagnóstico", "")], mc=[8])
            rn += 1
    pav_ag = res.get("pav_qr_agregado", pd.DataFrame())
    if pav_ag is not None and not pav_ag.empty:
        for _, a in pav_ag.iterrows():
            dw(ws_reg, rn, ["AGREGADO-PAVQR", "PAV QR agregado contra banco", a.get("Fecha Flexxus", ""), "PAV", a.get("Número Flexxus", ""), a.get("Concepto Flexxus", ""), a.get("Cantidad líneas banco", ""), float(a.get("Monto Flexxus", 0.0)), a.get("Diagnóstico", "")], mc=[8])
            rn += 1
    if rn == 3:
        ws_reg.cell(3, 1, "Sin regularizaciones en este período").font = norm
    aw(ws_reg)

    # HOJA 7: Carga Flexxus
    ws2 = wb.create_sheet("Carga Flexxus")
    ws2.cell(1, 1, "Movimientos para archivo de importación Flexxus").font = sub_f
    matched = res.get("matched_flex", pd.DataFrame()).copy()
    if matched is None:
        matched = pd.DataFrame()
    pav_match = matched[matched["Tipo"] == "PAV"] if not matched.empty and "Tipo" in matched.columns else pd.DataFrame()
    mbex_match = matched[matched["Tipo"] == "MB-ENT-EX"] if not matched.empty and "Tipo" in matched.columns else pd.DataFrame()
    mbext_match = matched[matched["Tipo"] == "MB-EXT"] if not matched.empty and "Tipo" in matched.columns else pd.DataFrame()
    total_qty = len(matched)
    total_amt = float(matched["MontoFlexxus"].sum()) if not matched.empty and "MontoFlexxus" in matched.columns else 0.0
    for i, (tipo, qty, total) in enumerate([
        ("PAV", len(pav_match), float(pav_match["MontoFlexxus"].sum()) if not pav_match.empty else 0.0),
        ("MB-ENT-EX", len(mbex_match), float(mbex_match["MontoFlexxus"].sum()) if not mbex_match.empty else 0.0),
        ("MB-EXT", len(mbext_match), float(mbext_match["MontoFlexxus"].sum()) if not mbext_match.empty else 0.0),
        ("TOTAL", total_qty, total_amt),
    ], 2):
        ws2.cell(i, 1, tipo).font = norm
        ws2.cell(i, 2, qty).font = norm
        c = ws2.cell(i, 3, total); c.number_format = money_fmt
    rn = 6
    hw(ws2, rn, ["Fecha mov. Flexxus", "Tipo", "Nro. Flexxus", "Monto", "Fecha banco real", "Fecha acreditación a importar", "Concepto banco", "Comprobante banco", "Ajuste fecha"])
    rn += 1
    if not matched.empty:
        for _, f in matched.iterrows():
            fa = f.get("FechaAcreditacionUsada", "") or f.get("BancoFecha", "") or f.get("FechaFlexxus", "")
            ajuste = "Sí" if f.get("BancoFecha", "") and f.get("BancoFecha", "") != fa else "No"
            dw(ws2, rn, [f.get("FechaFlexxus", ""), f.get("Tipo", ""), f.get("Numero", ""), float(f.get("MontoFlexxus", 0.0)), f.get("BancoFecha", ""), fa, f.get("BancoConcepto", ""), f.get("BancoComprobante", ""), ajuste], mc=[4])
            rn += 1
    frz(ws2); aw(ws2)

    # HOJA 8: Auditoría QR PCT
    ws7 = wb.create_sheet("Auditoría QR PCT")
    hw(ws7, 1, ["Fecha QR", "Estado", "Comercio", "Local", "Terminal", "Cupón/ID", "Monto Bruto", "Neto Calc.", "En Banco", "En Flexxus", "Estado Match"])
    rn = 2
    if qr is not None and not qr.empty:
        for _, q in qr.iterrows():
            cod = str(q.get("CodComercio", ""))
            local = LOCAL_MAP.get(cod, cod)
            terminal = q.get("Terminal", "")
            if terminal == "" and "Terminal" in q.index:
                terminal = q.get("Terminal", "")
            neto = float(q.get("NetoQR", 0.0))
            in_b = "Sí" if (not bank.empty and "Categoria" in bank.columns and len(bank[(bank["Categoria"] == "QR") & (abs(bank["ImporteAbs"] - neto) < 1.0)]) > 0) else "No"
            in_f = "Sí" if (not matched.empty and "MontoFlexxus" in matched.columns and len(matched[abs(matched["MontoFlexxus"] - neto) < 1.0]) > 0) else "No"
            estado_match = "OK - en banco y Flexxus" if in_b == "Sí" and in_f == "Sí" else ("En banco, NO en Flexxus" if in_b == "Sí" else "Revisar")
            dw(ws7, rn, [q.get("FechaQR", ""), q.get("Estado", ""), cod, local, terminal, q.get("Cupon", ""), float(q.get("MontoTotal", 0.0)), round(neto, 2), in_b, in_f, estado_match], mc=[7, 8])
            rn += 1
    frz(ws7); aw(ws7)

    # HOJA 9: Auditoría TRX Merchant
    ws8 = wb.create_sheet("Auditoría TRX Merchant")
    hw(ws8, 1, ["Fecha Pago", "Nro Liquidación", "Comercio", "Local", "Tarjeta", "Monto Neto", "Matcheado", "Estado"])
    rn = 2
    if trx is not None and not trx.empty:
        for _, t in trx.iterrows():
            comercio = str(t.get("COMERCIO", ""))
            local = LOCAL_MAP.get(comercio, t.get("Local", comercio))
            amt = float(t.get("MontoNeto", 0.0))
            in_m = "Sí" if (not matched.empty and "MontoFlexxus" in matched.columns and len(matched[abs(matched["MontoFlexxus"] - amt) < 1.0]) > 0) else "No"
            dw(ws8, rn, [t.get("FECHA DE PAGO", ""), t.get("NUMERO LIQUIDACION", ""), comercio, local, t.get("TARJETA", ""), amt, in_m, "OK" if in_m == "Sí" else "Revisar"], mc=[6])
            rn += 1
    frz(ws8); aw(ws8)

    # HOJA 10: Resumen atraso por local
    ws_loc = wb.create_sheet("Resumen atraso por local")
    ws_loc.cell(1, 1, "Resumen de atraso/carga pendiente por local").font = sub_f
    ws_loc.cell(2, 1, "Este tablero no sanciona por timing bancario normal; marca pendientes atribuibles o no identificados para seguimiento operativo.").font = norm
    local_rows = []

    if not pbi.empty:
        for _, p in pbi.iterrows():
            local, cupon, liq, monto_comp, dif, archivo_aux, diag_local = enrich_bank_income_pending_row(p)
            cat = p.get("Categoria", "")
            accion = "Identificar local/origen" if local == "No identificado" else "Revisar carga pendiente en Flexxus / cierre administrativo del local"
            add_local_delay_rows(local_rows, local, "Banco ingreso no Flexxus", cat, p.get("Monto", 0.0), p.get("Estado", ""), diag_local, accion)

    if not pf.empty:
        for _, p in pf.iterrows():
            local, cupon, liq, monto_comp, dif = enrich_flexxus_pending_row(p)
            loc = local if local not in ["", "—"] else "No identificado"
            cat = p.get("Categoria", p.get("TipoMovimiento", ""))
            add_local_delay_rows(local_rows, loc, "Flexxus no Banco", cat, p.get("Monto", 0.0), p.get("Estado", ""), "Detectado desde Flexxus/TRX/QR" if loc != "No identificado" else "Flexxus pendiente sin local detectado", "Verificar acreditación bancaria o fecha de cierre")

    if not pbe.empty:
        # Egresos bancarios/impuestos suelen ser administración central, pero se incluyen para no perder trazabilidad.
        for _, p in pbe.iterrows():
            cat = p.get("Categoria", "")
            local = "Administración/Banco" if cat in IMP_CATS else "No identificado"
            accion = flexxus_copy_instruction(cat, float(p.get("Monto", 0.0)), p.get("FechaOrigen", ""), p.get("Concepto", "")) if cat in IMP_CATS else "Revisar egreso bancario"
            add_local_delay_rows(local_rows, local, "Banco egreso no Flexxus", cat, p.get("Monto", 0.0), p.get("Estado", ""), "Banco / impuesto / gasto" if cat in IMP_CATS else "Banco egreso sin local", accion)

    hw(ws_loc, 4, ["Local", "Cantidad pendientes", "Importe total observado", "Banco ingresos no Flexxus", "Flexxus no Banco", "Banco egresos no Flexxus", "Categoría dominante", "Severidad", "Comunicado sugerido"])
    rn_loc = 5
    if local_rows:
        ldf = pd.DataFrame(local_rows)
        # Resumen por local
        for local, grp in ldf.groupby("Local", dropna=False):
            cantidad = int(grp["Cantidad"].sum())
            importe_total = float(grp["Importe"].sum())
            bi = float(grp.loc[grp["Bloque"].eq("Banco ingreso no Flexxus"), "Importe"].sum())
            fb = float(grp.loc[grp["Bloque"].eq("Flexxus no Banco"), "Importe"].sum())
            be = float(grp.loc[grp["Bloque"].eq("Banco egreso no Flexxus"), "Importe"].sum())
            cats = grp.groupby("Categoría")["Cantidad"].sum().sort_values(ascending=False)
            cat_dom = str(cats.index[0]) if not cats.empty else ""
            if local == "No identificado" or bi > 500000 or cantidad >= 10:
                sev = "ALTA"
            elif bi > 100000 or cantidad >= 4:
                sev = "MEDIA"
            else:
                sev = "BAJA"
            if local == "No identificado":
                comunicado = "Identificar origen/local de movimientos bancarios no registrados antes del cierre."
            elif bi > 0:
                comunicado = f"{local}: revisar carga pendiente en Flexxus por ingresos bancarios no registrados y cierre administrativo."
            elif fb > 0:
                comunicado = f"{local}: verificar PAV/ventas cargadas en Flexxus pendientes de acreditación bancaria."
            else:
                comunicado = f"{local}: revisar egresos/impuestos/gastos pendientes de registro."
            dw(ws_loc, rn_loc, [local, cantidad, importe_total, bi, fb, be, cat_dom, sev, comunicado], mc=[3,4,5,6])
            if sev == "ALTA":
                ws_loc.cell(rn_loc, 8).fill = red_fill
                ws_loc.cell(rn_loc, 8).font = Font(bold=True, color="C00000")
            elif sev == "MEDIA":
                ws_loc.cell(rn_loc, 8).fill = PatternFill("solid", fgColor="FFF2CC")
                ws_loc.cell(rn_loc, 8).font = Font(bold=True, color="9C6500")
            else:
                ws_loc.cell(rn_loc, 8).fill = grn_fill
                ws_loc.cell(rn_loc, 8).font = Font(bold=True, color="375623")
            rn_loc += 1

        rn_loc += 2
        ws_loc.cell(rn_loc, 1, "Detalle de hallazgos por local").font = sub_f
        hw(ws_loc, rn_loc + 1, ["Local", "Bloque", "Categoría", "Importe", "Estado trazabilidad", "Origen detectado", "Acción sugerida"])
        rn_loc += 2
        for _, r in ldf.sort_values(["Local", "Bloque", "Categoría"]).iterrows():
            dw(ws_loc, rn_loc, [r["Local"], r["Bloque"], r["Categoría"], float(r["Importe"]), r["Estado trazabilidad"], r["Origen detectado"], r["Acción sugerida"]], mc=[4])
            rn_loc += 1
    else:
        ws_loc.cell(rn_loc, 1, "Sin pendientes atribuibles a local en esta corrida.").font = norm
    frz(ws_loc); aw(ws_loc, mx=70)


    # HOJA 11: Control cobertura
    ws_cov = wb.create_sheet("Control cobertura")
    ws_cov.cell(1, 1, "Control de cobertura total: ninguna línea puede quedar en zona muerta").font = sub_f
    hw(ws_cov, 2, ["Fuente", "Control", "Total fuente", "Total clasificado", "Diferencia", "Estado", "Observación"])

    matched_bank = res.get("matched_bank", pd.DataFrame())
    consumed_bank = res.get("consumed_prev_bank", pd.DataFrame())
    matched_flex = res.get("matched_flex", pd.DataFrame())
    consumed_flex = res.get("consumed_prev_flex", pd.DataFrame())

    def df_sum(df, col, mask=None):
        try:
            if df is None or df.empty or col not in df.columns:
                return 0.0
            dd = df[mask] if mask is not None else df
            if dd is None or dd.empty:
                return 0.0
            return float(dd[col].sum())
        except Exception:
            return 0.0

    bank_ing_total = df_sum(bank, "ImporteAbs", bank["EsIngreso"] if (bank is not None and not bank.empty and "EsIngreso" in bank.columns) else None)
    bank_egr_total = df_sum(bank, "ImporteAbs", bank["EsEgreso"] if (bank is not None and not bank.empty and "EsEgreso" in bank.columns) else None)
    bank_ing_clas = df_sum(matched_bank, "ImporteAbs", matched_bank["EsIngreso"] if (matched_bank is not None and not matched_bank.empty and "EsIngreso" in matched_bank.columns) else None) + df_sum(consumed_bank, "ImporteAbs", consumed_bank["EsIngreso"] if (consumed_bank is not None and not consumed_bank.empty and "EsIngreso" in consumed_bank.columns) else None) + (float(pbi["Monto"].sum()) if pbi is not None and not pbi.empty else 0.0)
    bank_egr_clas = df_sum(matched_bank, "ImporteAbs", matched_bank["EsEgreso"] if (matched_bank is not None and not matched_bank.empty and "EsEgreso" in matched_bank.columns) else None) + df_sum(consumed_bank, "ImporteAbs", consumed_bank["EsEgreso"] if (consumed_bank is not None and not consumed_bank.empty and "EsEgreso" in consumed_bank.columns) else None) + (float(pbe["Monto"].sum()) if pbe is not None and not pbe.empty else 0.0)

    flex_pav_total = df_sum(flex, "MontoFlexxus", flex["Tipo"].astype(str).eq("PAV") if (flex is not None and not flex.empty and "Tipo" in flex.columns) else None)
    flex_egr_total = df_sum(flex, "MontoFlexxus", flex["Tipo"].astype(str).isin(["MB-ENT-EX", "MB-EXT"]) if (flex is not None and not flex.empty and "Tipo" in flex.columns) else None)
    flex_pav_clas = df_sum(matched_flex, "MontoFlexxus", matched_flex["Tipo"].astype(str).eq("PAV") if (matched_flex is not None and not matched_flex.empty and "Tipo" in matched_flex.columns) else None) + df_sum(consumed_flex, "MontoFlexxus", consumed_flex["Tipo"].astype(str).eq("PAV") if (consumed_flex is not None and not consumed_flex.empty and "Tipo" in consumed_flex.columns) else None) + (float(pf[pf["TipoMovimiento"].astype(str).eq("PAV")]["Monto"].sum()) if pf is not None and not pf.empty else 0.0)
    flex_egr_clas = df_sum(matched_flex, "MontoFlexxus", matched_flex["Tipo"].astype(str).isin(["MB-ENT-EX", "MB-EXT"]) if (matched_flex is not None and not matched_flex.empty and "Tipo" in matched_flex.columns) else None) + df_sum(consumed_flex, "MontoFlexxus", consumed_flex["Tipo"].astype(str).isin(["MB-ENT-EX", "MB-EXT"]) if (consumed_flex is not None and not consumed_flex.empty and "Tipo" in consumed_flex.columns) else None) + (float(pf[~pf["TipoMovimiento"].astype(str).eq("PAV")]["Monto"].sum()) if pf is not None and not pf.empty else 0.0)

    cov_rows = [
        ("Banco", "Ingresos banco: real = conciliado + regularizado + Banco ingresos no Flexxus", bank_ing_total, bank_ing_clas),
        ("Banco", "Egresos banco: real = conciliado + regularizado + Banco egresos no Flexxus", bank_egr_total, bank_egr_clas),
        ("Flexxus", "PAV Flexxus: real = conciliado + regularizado + Flexxus no Banco", flex_pav_total, flex_pav_clas),
        ("Flexxus", "Egresos Flexxus: real = conciliado + regularizado + Flexxus egresos no Banco", flex_egr_total, flex_egr_clas),
    ]
    rn_cov = 3
    for fuente, control, total, clas in cov_rows:
        diff = round(float(total) - float(clas), 2)
        estado = "OK" if abs(diff) <= 1.0 else "ERROR"
        obs = "Sin movimientos en zona muerta" if estado == "OK" else "Hay movimientos clasificados de forma incompleta; revisar detalle"
        dw(ws_cov, rn_cov, [fuente, control, total, clas, diff, estado, obs], mc=[3,4,5])
        if estado == "ERROR":
            ws_cov.cell(rn_cov, 6).fill = red_fill
            ws_cov.cell(rn_cov, 6).font = Font(bold=True, color="C00000")
        else:
            ws_cov.cell(rn_cov, 6).fill = grn_fill
            ws_cov.cell(rn_cov, 6).font = Font(bold=True, color="375623")
        rn_cov += 1

    rn_cov += 2
    ws_cov.cell(rn_cov, 1, "Movimientos bancarios pendientes visibles").font = sub_f
    hw(ws_cov, rn_cov + 1, ["Fuente", "Fecha", "Comprobante/Número", "Concepto", "Importe", "Estado interno", "Diagnóstico"])
    rn_cov += 2
    try:
        bank_unclassified = bank[(~bank["Matched"]) & (~bank["ConsumedPrev"])].copy()
        if not bank_unclassified.empty:
            for _, b in bank_unclassified.iterrows():
                dw(ws_cov, rn_cov, ["Banco", b.get("Fecha", ""), b.get("Comprobante", ""), b.get("Concepto", ""), float(b.get("ImporteAbs", 0.0)), "Pendiente visible en hojas Banco no Flexxus", b.get("Diagnostico", "")], mc=[5])
                rn_cov += 1
    except Exception:
        pass
    frz(ws_cov); aw(ws_cov, mx=70)

    # HOJA 10: Notas proceso
    ws_n = wb.create_sheet("Notas proceso")
    hw(ws_n, 1, ["Tema", "Detalle"])
    notas = [
        ("Regla base", "Conciliación de Flexxus define movimientos pendientes, tipo y fecha de movimiento. Banco confirma acreditación/débito real."),
        ("Regla de carga", "Solo se carga en Flexxus si existe en banco. Merchant/QR se usa como auditoría auxiliar, no como base directa de carga."),
        ("QR Banco Nación", "Todo CR DEBIN SPOT se considera QR/PAV."),
        ("QR Merchant", "Neto QR = Monto total x 0,99032; se usa para identificar local y cupón cuando corresponde."),
        ("Local", "Locales: " + ", ".join([f"{k}={v}" for k, v in LOCAL_MAP.items()])),
        ("Pedidos Ya", "PEDIDOS YA es cliente/canal aparte, no tarjeta ni QR; no completar cupón QR ni liquidación."),
        ("Fechas Flexxus", "Si fecha banco < fecha Flexxus, el archivo final usa FECHAACREDITACION = FECHAMOVIMIENTO; la fecha real queda en auditoría."),
        ("V5.8", "Versión final: separa saldo Flexxus real, saldo Flexxus ajustado para conciliación y total importado .xls; mantiene fórmulas reales y control de cobertura."),
        ("V5.6", "Separa pendientes reales abiertos de movimientos ya procesados/históricos en la pantalla principal y en el conteo del historial."),
        ("V5.5", "Corrige identificación de local/liquidación en Banco ingresos no Flexxus: usa TRX exacto, QR PCT y fallback por comprobante banco=código de comercio."),
    ]
    for i, (t, d) in enumerate(notas, 2):
        ws_n.cell(i, 1, t).font = bold_f
        ws_n.cell(i, 2, d).font = norm
    ws_n.column_dimensions["A"].width = 25
    ws_n.column_dimensions["B"].width = 90

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out

def build_import_xls(res: Dict) -> io.BytesIO:
    wb = xlwt.Workbook()
    ws = wb.add_sheet("ASIENTOS")
    header_style = xlwt.easyxf("font: bold on; align: horiz center")
    date_style = xlwt.easyxf(num_format_str="DD/MM/YYYY")
    money_style = xlwt.easyxf(num_format_str="#,##0.00")
    headers = ["FECHAMOVIMIENTO", "TIPOMOVIMIENTO", "MONTO", "FECHAACREDITACION"]
    for c, h in enumerate(headers):
        ws.write(0, c, h, header_style)
    matched = res["matched_flex"].copy()
    # Solo movimientos corrientes. No incluye ConsumedPrev.
    for i, (_, fr) in enumerate(matched.iterrows(), 1):
        fm = safe_date(fr["FechaFlexxus"])
        fa = safe_date(fr.get("FechaAcreditacionUsada", "") or fr.get("BancoFecha", "") or fr["FechaFlexxus"])
        if pd.notna(fm):
            ws.write(i, 0, fm.to_pydatetime(), date_style)
        else:
            ws.write(i, 0, str(fr["FechaFlexxus"]))
        ws.write(i, 1, str(fr["Tipo"]))
        ws.write(i, 2, float(abs(fr["MontoFlexxus"])), money_style)
        if pd.notna(fa):
            ws.write(i, 3, fa.to_pydatetime(), date_style)
        else:
            ws.write(i, 3, str(fr.get("FechaAcreditacionUsada", "")))
    for c in range(4):
        ws.col(c).width = 5000
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out


# =============================================================================
# UI STREAMLIT V4 INTEGRADA
# =============================================================================

APP_VERSION = "V5.8-FINAL-SALDOS-SEPARADOS-FORMULAS-COBERTURA-2026-04-30"
HISTORIAL_FILE = "historico.json"

def secret_get(key: str, default: str = "") -> str:
    try:
        return st.secrets.get(key, default)
    except Exception:
        return default

def render_header_v4():
    st.markdown(f"""
    <div style="background:linear-gradient(135deg,#1F4E78,#0F243E);padding:28px;border-radius:14px;margin-bottom:20px;color:white">
      <div style="font-size:13px;opacity:.75;letter-spacing:.08em">GRUPO DANCONA · CONTROL BANCARIO</div>
      <div style="font-size:30px;font-weight:700">Conciliación Bancaria Continua V5.8</div>
      <div style="font-size:15px;opacity:.85;margin-top:6px">Login · Historial con archivos · Administrador · Botón Comenzar · Pendientes anteriores · Matching agregado MB-EXT · Sin ajustes genéricos.</div>
      <div style="font-size:11px;opacity:.70;margin-top:10px;font-family:monospace">{APP_VERSION}</div>
    </div>
    """, unsafe_allow_html=True)

def login_gate():
    if "authenticated" not in st.session_state:
        st.session_state.authenticated = False
    if st.session_state.authenticated:
        return
    render_header_v4()
    st.subheader("🔐 Iniciar sesión")
    valid_user = secret_get("APP_USER", "dancona2016@gmail.com")
    valid_password = secret_get("APP_PASSWORD", "Dancona2026*")
    col1, col2, col3 = st.columns([1, 1.4, 1])
    with col2:
        usuario = st.text_input("Usuario", placeholder="email")
        password = st.text_input("Contraseña", type="password")
        if st.button("Ingresar", type="primary", use_container_width=True):
            if usuario == valid_user and password == valid_password:
                st.session_state.authenticated = True
                st.rerun()
            else:
                st.error("Usuario o contraseña incorrectos.")
    st.stop()

def reset_app_state():
    keep = {"authenticated"}
    for k in list(st.session_state.keys()):
        if k not in keep:
            del st.session_state[k]
    try:
        st.cache_data.clear()
        st.cache_resource.clear()
    except Exception:
        pass

def github_headers():
    token = secret_get("GITHUB_TOKEN", "")
    return {"Authorization": f"token {token}", "Accept": "application/vnd.github.v3+json"}

def github_repo_ready():
    token = secret_get("GITHUB_TOKEN", "")
    repo = secret_get("GITHUB_REPO", "")
    return bool(token and repo and requests is not None), repo

def github_get_file(filename: str = HISTORIAL_FILE):
    ready, repo = github_repo_ready()
    if not ready:
        return None, None, "GitHub no configurado o requests no disponible"
    url = f"https://api.github.com/repos/{repo}/contents/{filename}"
    try:
        r = requests.get(url, headers=github_headers(), timeout=15)
        if r.status_code == 200:
            data = r.json()
            content = json.loads(base64.b64decode(data["content"]).decode("utf-8"))
            return content, data.get("sha"), "OK"
        if r.status_code == 404:
            return {"semanas": []}, None, "Historial no existe todavía; se creará al guardar."
        return None, None, f"GitHub HTTP {r.status_code}: {r.text[:250]}"
    except Exception as e:
        return None, None, f"Error leyendo GitHub: {e}"

def github_save_file(content_dict: dict, sha=None, filename: str = HISTORIAL_FILE):
    ready, repo = github_repo_ready()
    if not ready:
        return False, "GitHub no configurado. La conciliación se generó, pero no se guardó historial."
    url = f"https://api.github.com/repos/{repo}/contents/{filename}"
    encoded = base64.b64encode(json.dumps(content_dict, ensure_ascii=False, indent=2).encode("utf-8")).decode("utf-8")
    body = {"message": f"Update {filename} - conciliacion V4.7", "content": encoded}
    if sha:
        body["sha"] = sha
    try:
        r = requests.put(url, headers=github_headers(), json=body, timeout=20)
        if r.status_code in (200, 201):
            return True, "Historial guardado en GitHub."
        return False, f"GitHub HTTP {r.status_code}: {r.text[:300]}"
    except Exception as e:
        return False, f"Error guardando GitHub: {e}"

def github_get_content_meta(path: str):
    ready, repo = github_repo_ready()
    if not ready:
        return None, "GitHub no configurado"
    url = f"https://api.github.com/repos/{repo}/contents/{path}"
    try:
        r = requests.get(url, headers=github_headers(), timeout=20)
        if r.status_code == 200:
            return r.json(), "OK"
        if r.status_code == 404:
            return None, "No existe"
        return None, f"GitHub HTTP {r.status_code}: {r.text[:250]}"
    except Exception as e:
        return None, f"Error leyendo {path}: {e}"

def github_save_bytes(path: str, data: bytes, message: str):
    ready, repo = github_repo_ready()
    if not ready:
        return False, "GitHub no configurado"
    meta, _ = github_get_content_meta(path)
    url = f"https://api.github.com/repos/{repo}/contents/{path}"
    body = {"message": message, "content": base64.b64encode(data).decode("utf-8")}
    if meta and meta.get("sha"):
        body["sha"] = meta["sha"]
    try:
        r = requests.put(url, headers=github_headers(), json=body, timeout=30)
        if r.status_code in (200, 201):
            return True, "OK"
        return False, f"GitHub HTTP {r.status_code}: {r.text[:300]}"
    except Exception as e:
        return False, f"Error guardando archivo {path}: {e}"

def github_get_bytes(path: str):
    meta, msg = github_get_content_meta(path)
    if not meta:
        return None, msg
    try:
        return base64.b64decode(meta["content"]), "OK"
    except Exception as e:
        return None, f"Error decodificando {path}: {e}"

def github_delete_file(path: str, message: str):
    ready, repo = github_repo_ready()
    if not ready:
        return False, "GitHub no configurado"
    meta, msg = github_get_content_meta(path)
    if not meta:
        return True, f"Archivo no encontrado, se considera eliminado: {path}"
    url = f"https://api.github.com/repos/{repo}/contents/{path}"
    body = {"message": message, "sha": meta.get("sha")}
    try:
        r = requests.delete(url, headers=github_headers(), json=body, timeout=20)
        if r.status_code == 200:
            return True, "OK"
        return False, f"GitHub HTTP {r.status_code}: {r.text[:300]}"
    except Exception as e:
        return False, f"Error eliminando {path}: {e}"

def guardar_resumen_historial(res: Dict, mode: str, xlsx_bytes: bytes = None, xls_bytes: bytes = None):
    hist, sha, msg = github_get_file()
    if hist is None:
        return False, msg
    hist.setdefault("semanas", [])
    item_id = datetime.now().strftime("%Y%m%d_%H%M%S")
    archivos = []
    if xlsx_bytes:
        path_xlsx = f"historial/{item_id}/Conciliacion_Semanal_Dancona_{item_id}.xlsx"
        ok_file, msg_file = github_save_bytes(path_xlsx, xlsx_bytes, f"Guardar conciliación Excel {item_id}")
        if ok_file:
            archivos.append({"tipo": "conciliacion_xlsx", "nombre": f"Conciliacion_Semanal_Dancona_{item_id}.xlsx", "path": path_xlsx, "bytes": len(xlsx_bytes)})
        else:
            archivos.append({"tipo": "error_guardado_xlsx", "nombre": "ERROR", "path": path_xlsx, "error": msg_file})
    if xls_bytes:
        path_xls = f"historial/{item_id}/ASIENTOS_FLEXXUS_{item_id}.xls"
        ok_file, msg_file = github_save_bytes(path_xls, xls_bytes, f"Guardar import Flexxus {item_id}")
        if ok_file:
            archivos.append({"tipo": "import_flexxus_xls", "nombre": f"ASIENTOS_FLEXXUS_{item_id}.xls", "path": path_xls, "bytes": len(xls_bytes)})
        else:
            archivos.append({"tipo": "error_guardado_xls", "nombre": "ERROR", "path": path_xls, "error": msg_file})
    item = {
        "id": item_id,
        "fecha_proceso": datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
        "version_app": APP_VERSION,
        "modo": mode,
        "saldo_flexxus": round(float(res.get("saldo_flexxus", 0)), 2),
        "saldo_banco": round(float(res.get("saldo_banco", 0)), 2),
        "banco_calculado": round(float(res.get("calc_final", 0)), 2),
        "diferencia": round(float(res.get("diferencia", 0)), 2),
        "C1_flexxus_ingresos_no_banco": round(float(res.get("C1", 0)), 2),
        "C2_flexxus_egresos_no_banco": round(float(res.get("C2", 0)), 2),
        "C3_banco_ingresos_no_flexxus": round(float(res.get("C3", 0)), 2),
        "C4_banco_egresos_no_flexxus": round(float(res.get("C4", 0)), 2),
        "pendientes_anteriores_abiertos": int(len(res.get("prev_open", pd.DataFrame()))),
        "regularizaciones_anteriores": int(len(res.get("regularizaciones", pd.DataFrame()))),
        "pendientes_proxima": int(len(split_pendientes_para_ui(res.get("pendientes_proxima", pd.DataFrame()))[0])),
        "trazabilidad_procesada_anterior": int(len(split_pendientes_para_ui(res.get("pendientes_proxima", pd.DataFrame()))[1])),
        "mbext_agregados": int(len(res.get("mbext_agregado", pd.DataFrame()))),
        "archivos": archivos,
        "eliminado": False,
    }
    hist["semanas"].append(item)
    ok, save_msg = github_save_file(hist, sha)
    if not ok:
        return False, save_msg
    if any(str(a.get("tipo", "")).startswith("error_guardado") for a in archivos):
        return True, "Resumen guardado, pero hubo errores guardando uno o más archivos. Revisá el historial."
    return True, "Historial y archivos guardados en GitHub."

def eliminar_item_historial(item_id: str, borrar_archivos: bool = True):
    hist, sha, msg = github_get_file()
    if hist is None:
        return False, msg
    semanas = hist.get("semanas", [])
    item = next((x for x in semanas if str(x.get("id")) == str(item_id)), None)
    if not item:
        return False, "No encontré ese registro en historico.json."
    if borrar_archivos:
        for a in item.get("archivos", []):
            path = a.get("path")
            if path:
                ok_del, msg_del = github_delete_file(path, f"Eliminar archivo historial {item_id}: {path}")
                if not ok_del:
                    return False, "No se pudo eliminar un archivo: " + msg_del
    hist["semanas"] = [x for x in semanas if str(x.get("id")) != str(item_id)]
    ok, save_msg = github_save_file(hist, sha)
    if not ok:
        return False, save_msg
    return True, "Registro eliminado del historial" + (" y archivos eliminados." if borrar_archivos else ".")

def render_historial_tab():
    st.subheader("📊 Historial")
    hist, sha, msg = github_get_file()
    if hist is None:
        st.warning(msg)
        st.info("La app puede conciliar igual. Para guardar historial con archivos configurá GITHUB_TOKEN y GITHUB_REPO en Secrets.")
        return
    semanas = hist.get("semanas", [])
    if not semanas:
        st.info("Todavía no hay conciliaciones guardadas en el historial.")
        st.caption(msg)
        return
    df = pd.DataFrame(semanas)
    if "fecha_proceso" in df.columns:
        df = df.sort_values("fecha_proceso", ascending=False)
    display_cols = [c for c in ["id", "fecha_proceso", "version_app", "modo", "saldo_flexxus", "saldo_banco", "banco_calculado", "diferencia", "pendientes_proxima", "regularizaciones_anteriores", "mbext_agregados"] if c in df.columns]
    st.dataframe(df[display_cols] if display_cols else df, use_container_width=True, hide_index=True)
    st.download_button("Descargar historico.json", data=json.dumps(hist, ensure_ascii=False, indent=2).encode("utf-8"), file_name="historico.json", mime="application/json", use_container_width=True)

    ids = [str(x.get("id")) for x in semanas if x.get("id")]
    if not ids:
        return
    st.markdown("---")
    st.subheader("🗂️ Archivos del historial")
    selected_id = st.selectbox("Seleccionar conciliación", options=ids, index=len(ids)-1 if ids else 0)
    selected = next((x for x in semanas if str(x.get("id")) == str(selected_id)), None)
    if selected:
        st.json({k: v for k, v in selected.items() if k != "archivos"})
        archivos = selected.get("archivos", [])
        if archivos:
            st.write("Archivos guardados")
            st.dataframe(pd.DataFrame(archivos), use_container_width=True, hide_index=True)
            for a in archivos:
                path = a.get("path", "")
                nombre = a.get("nombre", path.split("/")[-1])
                tipo = a.get("tipo", "archivo")
                if path and not str(tipo).startswith("error"):
                    data, msg_b = github_get_bytes(path)
                    if data:
                        mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" if nombre.lower().endswith(".xlsx") else "application/vnd.ms-excel"
                        st.download_button(f"Descargar {nombre}", data=data, file_name=nombre, mime=mime, use_container_width=True, key=f"dl_{selected_id}_{path}")
                    else:
                        st.warning(f"No pude leer {path}: {msg_b}")
        else:
            st.info("Este registro no tiene archivos asociados. Probablemente fue guardado con una versión anterior del historial.")

    with st.expander("🛡️ Administrador del historial"):
        st.warning("Eliminar borra el registro del historico.json y, si existen, los archivos Excel/XLS guardados en GitHub para esa conciliación. No borra archivos locales de tu computadora.")
        confirm_txt = st.text_input("Para eliminar, escribí ELIMINAR", key="confirm_delete_hist")
        borrar_archivos = st.checkbox("Eliminar también archivos asociados", value=True)
        if st.button("Eliminar conciliación seleccionada", type="secondary", use_container_width=True):
            if confirm_txt.strip().upper() != "ELIMINAR":
                st.error("No se eliminó. Tenés que escribir ELIMINAR.")
            else:
                ok, del_msg = eliminar_item_historial(selected_id, borrar_archivos=borrar_archivos)
                if ok:
                    st.success(del_msg)
                    st.rerun()
                else:
                    st.error(del_msg)

def render_diagnostico_tab():
    st.subheader("🧪 Diagnóstico técnico")
    st.code(APP_VERSION)
    st.write("Python")
    st.code(sys.version)
    versions = {
        "streamlit": getattr(st, "__version__", ""),
        "pandas": pd.__version__,
        "numpy": np.__version__,
        "requests": getattr(requests, "__version__", "NO DISPONIBLE") if requests is not None else "NO DISPONIBLE",
        "xlwt": getattr(xlwt, "__VERSION__", "instalado"),
    }
    for mod in ["openpyxl", "xlrd"]:
        try:
            m = __import__(mod)
            versions[mod] = getattr(m, "__version__", "instalado")
        except Exception as e:
            versions[mod] = f"ERROR: {e}"
    st.json(versions)
    st.write("Estado de sesión")
    st.json({k: str(v)[:120] for k, v in st.session_state.items() if k not in {"password"}})
    st.write("Secrets detectados")
    st.json({"APP_USER": bool(secret_get("APP_USER", "")), "APP_PASSWORD": bool(secret_get("APP_PASSWORD", "")), "GITHUB_TOKEN": bool(secret_get("GITHUB_TOKEN", "")), "GITHUB_REPO": secret_get("GITHUB_REPO", "")})

def render_conciliacion_tab():
    st.subheader("🏦 Conciliación")
    st.info("Subí Flexxus y Banco como mínimo. TRX, QR y conciliación anterior mejoran la auditoría. Desde la segunda semana, subí siempre la conciliación anterior para cancelar pendientes.")
    col_a, col_b = st.columns(2)
    with col_a:
        f_flex = st.file_uploader("1 · Conciliación Bancaria Flexxus actual", type=["xls", "xlsx"], key="f_flex_v4")
        f_trx = st.file_uploader("3 · TRX Merchant Center", type=["xlsx", "xls"], key="f_trx_v4")
        f_prev = st.file_uploader("5 · Conciliación anterior del sistema", type=["xlsx"], key="f_prev_v4")
    with col_b:
        f_bank = st.file_uploader("2 · Últimos movimientos Banco Nación actual", type=["xls", "xlsx"], key="f_bank_v4")
        f_qr = st.file_uploader("4 · Transacciones QR", type=["xlsx", "xls"], key="f_qr_v4")
    with st.expander("🧪 Diagnóstico de archivos cargados"):
        files = {"Flexxus": f_flex, "Banco": f_bank, "TRX": f_trx, "QR": f_qr, "Conciliación anterior": f_prev}
        st.dataframe(pd.DataFrame([{"Archivo": k, "Cargado": v is not None, "Nombre": getattr(v, "name", ""), "Tamaño bytes": getattr(v, "size", 0) if v is not None else 0} for k, v in files.items()]), use_container_width=True, hide_index=True)

    if not f_flex or not f_bank:
        st.warning("Subí al menos Flexxus actual y Banco Nación actual para habilitar el procesamiento.")
        return

    run = st.button("▶️ Comenzar conciliación", type="primary", use_container_width=True)
    if st.button("🧹 Limpiar resultado", use_container_width=True):
        for k in ["last_result_v4", "last_error_v4", "last_xlsx_v4", "last_xls_v4"]:
            st.session_state.pop(k, None)
        st.rerun()

    if run:
        try:
            with st.spinner("Leyendo archivos..."):
                flex = parse_flexxus(f_flex)
                bank = parse_banco(f_bank)
                trx = parse_trx(f_trx) if f_trx else pd.DataFrame()
                qr = parse_qr(f_qr) if f_qr else pd.DataFrame()
                prev_open, prev_summary, mode = parse_previous_conciliation(f_prev)
            with st.spinner("Cancelando pendientes anteriores contra movimientos actuales..."):
                prev_status, flex, bank, regs = match_previous_pendings(prev_open, flex, bank, prev_summary)
            with st.spinner("Conciliando movimientos corrientes..."):
                flex, bank = match_current(flex, bank, qr, trx)
                res = compute_results(flex, bank, prev_status, regs, mode, prev_summary)
            with st.spinner("Generando entregables..."):
                xlsx = build_excel_report(flex, bank, qr, trx, res)
                xls = build_import_xls(res)
            st.session_state.last_result_v4 = res
            st.session_state.last_xlsx_v4 = xlsx.getvalue()
            st.session_state.last_xls_v4 = xls.getvalue()
            st.session_state.last_error_v4 = None
            st.success("Conciliación procesada.")
        except Exception as e:
            err = traceback.format_exc()
            st.session_state.last_error_v4 = err
            st.exception(e)
            st.error("No se pudo procesar. Descargá el diagnóstico y pasámelo.")
            st.download_button("Descargar diagnóstico del error", data=err.encode("utf-8"), file_name="diagnostico_error_v4.txt", mime="text/plain")
            return

    if "last_result_v4" not in st.session_state:
        st.caption("Esperando que presiones **Comenzar conciliación**.")
        return

    res = st.session_state.last_result_v4
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Saldo Flexxus", f"${res['saldo_flexxus']:,.2f}")
    c2.metric("Saldo Banco", f"${res['saldo_banco']:,.2f}")
    c3.metric("Banco calculado", f"${res['calc_final']:,.2f}")
    c4.metric("Diferencia", f"${res['diferencia']:,.2f}")

    cont = res.get("continuity", {})
    if cont.get("aplica"):
        st.subheader("🔗 Trazabilidad con semana anterior")
        cc1, cc2, cc3 = st.columns(3)
        cc1.metric("Saldo banco cierre anterior", f"${cont.get('saldo_banco_cierre_anterior',0):,.2f}")
        cc2.metric("Apertura banco actual", f"${cont.get('saldo_banco_apertura_actual',0):,.2f}")
        cc3.metric("Diferencia continuidad banco", f"${cont.get('diferencia_continuidad_banco',0):,.2f}")
        if cont.get("ok"):
            st.success(cont.get("mensaje"))
        else:
            st.warning(cont.get("mensaje"))
    else:
        st.info(cont.get("mensaje", "Sin control de continuidad bancaria."))

    st.subheader("Prueba explícita")
    proof = pd.DataFrame([
        {"Concepto": "Saldo Flexxus actual", "Importe": res["saldo_flexxus"]},
        {"Concepto": "- C1 corriente: ingresos Flexxus no Banco", "Importe": -res["C1"]},
        {"Concepto": "+ C2 corriente: egresos Flexxus no Banco", "Importe": res["C2"]},
        {"Concepto": "+ C3 corriente: ingresos Banco no Flexxus", "Importe": res["C3"]},
        {"Concepto": "- C4 corriente: egresos Banco no Flexxus", "Importe": -res["C4"]},
        {"Concepto": "+/- pendientes anteriores aún abiertos", "Importe": res["prev_open_effect"]},
        {"Concepto": "Banco calculado", "Importe": res["calc_final"]},
        {"Concepto": "Banco real", "Importe": res["saldo_banco"]},
        {"Concepto": "Diferencia final", "Importe": res["diferencia"]},
    ])
    st.dataframe(proof, use_container_width=True, hide_index=True)
    if abs(res["diferencia"]) >= 1:
        st.error("La diferencia final no da 0. El sistema NO la forzó. Revisá pendientes abiertos y movimientos no conciliados.")
    else:
        st.success("Diferencia final conciliada en 0,00 o dentro de tolerancia de redondeo.")

    st.subheader("Regularizaciones de períodos anteriores")
    if res["regularizaciones"].empty:
        st.write("Sin regularizaciones anteriores detectadas.")
    else:
        st.dataframe(res["regularizaciones"], use_container_width=True, hide_index=True)

    st.subheader("MB-EXT agregados contra banco")
    ag = res.get("mbext_agregado", pd.DataFrame())
    if ag.empty:
        st.write("Sin MB-EXT agregados en esta corrida.")
    else:
        st.dataframe(ag, use_container_width=True, hide_index=True)

    st.subheader("Pendientes abiertos reales para próxima conciliación")
    pend = res.get("pendientes_proxima", pd.DataFrame())
    pend_reales, pend_procesados, pend_controles = split_pendientes_para_ui(pend)

    pcol1, pcol2, pcol3 = st.columns(3)
    pcol1.metric("Pendientes reales abiertos", len(pend_reales))
    pcol2.metric("Trazabilidad procesada anterior", len(pend_procesados))
    pcol3.metric("Controles técnicos visibles", len(pend_controles))

    if pend_reales.empty:
        st.success("No quedan pendientes reales abiertos para próxima conciliación.")
        st.dataframe(pd.DataFrame([{"Estado": "Sin pendientes abiertos reales"}]), use_container_width=True, hide_index=True)
    else:
        st.dataframe(pend_reales, use_container_width=True, hide_index=True)

    if not pend_procesados.empty:
        with st.expander("Ver trazabilidad histórica / movimientos ya procesados (no son pendientes abiertos)", expanded=False):
            st.info("Estas filas se conservan para auditoría y reprocesos, pero no deben tratarse como pendientes abiertos.")
            st.dataframe(pend_procesados, use_container_width=True, hide_index=True)

    if not pend_controles.empty:
        with st.expander("Ver controles técnicos menores", expanded=False):
            st.dataframe(pend_controles, use_container_width=True, hide_index=True)

    d1, d2 = st.columns(2)
    with d1:
        st.download_button("Descargar Excel de conciliación V5.8", data=st.session_state.last_xlsx_v4, file_name="Conciliacion_Semanal_Dancona_V5_8.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
    with d2:
        st.download_button("Descargar .xls para importar a Flexxus", data=st.session_state.last_xls_v4, file_name="ASIENTOS_FLEXXUS_V5_8.xls", mime="application/vnd.ms-excel", use_container_width=True)

    if st.button("💾 Guardar resumen en historial", use_container_width=True):
        ok, msg = guardar_resumen_historial(res, res.get("mode", ""), st.session_state.get("last_xlsx_v4"), st.session_state.get("last_xls_v4"))
        if ok:
            st.success(msg)
        else:
            st.warning(msg)

def main():
    login_gate()
    with st.sidebar:
        st.markdown("### Sesión")
        if st.button("Cerrar sesión", use_container_width=True):
            st.session_state.authenticated = False
            reset_app_state()
            st.rerun()
        if st.button("🔄 Resetear app / limpiar estado", use_container_width=True):
            reset_app_state()
            st.rerun()
        st.caption(APP_VERSION)
    render_header_v4()
    tab1, tab2, tab3 = st.tabs(["🏦 Conciliación", "📊 Historial", "🧪 Diagnóstico"])
    with tab1:
        render_conciliacion_tab()
    with tab2:
        render_historial_tab()
    with tab3:
        render_diagnostico_tab()

if __name__ == "__main__":
    main()
